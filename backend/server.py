"""
Super Apps MATSANDATAMA - Main FastAPI Server
MTsN 2 Kota Malang - Sistem Jurnal Presisi Multi-Role
"""
import os
import io
import logging
import base64
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Optional, Dict, Any

from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Request, Query
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
import pyotp

from models import (
    ROLES, ROLE_LABELS,
    UserModel, AcademicYearModel, ClassModel, RoomModel, SubjectModel,
    ScheduleModel, JournalModel, AuditLogModel, SecurityLogModel,
    SettingsModel, QRTemplateModel,
    LoginRequest, LoginResponse, CaptchaResponse, RoleSwitchRequest,
    UserCreateRequest, UserUpdateRequest, JournalCreateRequest, QRValidateRequest,
    ClassAttendanceModel, ClassCleanlinessModel,
    ClassAttendanceSubmit, ClassCleanlinessSubmit,
)
from auth_utils import (
    hash_password, verify_password, create_access_token, decode_token,
    generate_math_captcha, verify_captcha,
    record_login_attempt, is_locked,
)
from journal_core import (
    encrypt_qr_payload, decrypt_qr_payload, generate_qr_image_b64,
    validate_gps, validate_schedule, create_b5_card,
    generate_dynamic_qr_payload, validate_dynamic_qr,
    now_wib, current_day_id, WIB_TZ,
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Super Apps MATSANDATAMA API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ============================================================
# HELPERS
# ============================================================
def serialize_doc(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Convert MongoDB doc into JSON-serializable dict."""
    if not doc:
        return doc
    doc.pop('_id', None)
    for k, v in list(doc.items()):
        if isinstance(v, datetime):
            doc[k] = v.isoformat()
        elif isinstance(v, dict):
            doc[k] = serialize_doc(v)
        elif isinstance(v, list):
            doc[k] = [serialize_doc(i) if isinstance(i, dict) else (i.isoformat() if isinstance(i, datetime) else i) for i in v]
    return doc


async def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Dict[str, Any]:
    if not credentials:
        raise HTTPException(status_code=401, detail="Tidak terautentikasi. Silakan login.")
    payload = decode_token(credentials.credentials)
    if not payload:
        raise HTTPException(status_code=401, detail="Token tidak valid atau kedaluwarsa")
    user_id = payload.get('sub')
    user = await db.users.find_one({'id': user_id})
    if not user or not user.get('is_active', True):
        raise HTTPException(status_code=401, detail="User tidak ditemukan atau dinonaktifkan")
    user['active_role'] = payload.get('active_role', user['roles'][0] if user.get('roles') else 'guru')
    return serialize_doc(user)


def require_role(*allowed_roles: str):
    async def checker(user: Dict[str, Any] = Depends(get_current_user)):
        active = user.get('active_role')
        if active not in allowed_roles and 'admin' not in user.get('roles', []):
            raise HTTPException(status_code=403, detail=f"Akses ditolak. Peran aktif: {active}")
        return user
    return checker


async def log_audit(user, action, entity, entity_id=None, details=None, request=None):
    log = AuditLogModel(
        user_id=user.get('id') if user else None,
        username=user.get('username') if user else None,
        action=action, entity=entity, entity_id=entity_id,
        details=details or {},
        ip_address=request.client.host if request else None,
        user_agent=request.headers.get('user-agent') if request else None,
    )
    doc = log.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.audit_logs.insert_one(doc)


async def log_security(event_type, username=None, details=None, request=None):
    log = SecurityLogModel(
        event_type=event_type, username=username,
        ip_address=request.client.host if request else None,
        user_agent=request.headers.get('user-agent') if request else None,
        details=details or {},
    )
    doc = log.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.security_logs.insert_one(doc)


async def get_active_academic_year():
    ay = await db.academic_years.find_one({'is_active': True})
    return serialize_doc(ay) if ay else None


async def get_settings():
    s = await db.settings.find_one({'id': 'global_config'})
    if not s:
        default = SettingsModel().model_dump()
        default['updated_at'] = default['updated_at'].isoformat()
        await db.settings.insert_one(default)
        return default
    return serialize_doc(s)


# ============================================================
# HEALTH & ROOT
# ============================================================
@api_router.get("/")
async def root():
    return {"message": "Super Apps MATSANDATAMA API", "status": "ok"}


@api_router.get("/health")
async def health():
    return {"status": "healthy", "time_wib": now_wib().isoformat()}


# ============================================================
# AUTH
# ============================================================
@api_router.get("/auth/captcha", response_model=CaptchaResponse)
async def get_captcha():
    return generate_math_captcha()


@api_router.post("/auth/login", response_model=LoginResponse)
async def login(req: LoginRequest, request: Request):
    if is_locked(req.username):
        await log_security('locked_attempt', req.username, request=request)
        raise HTTPException(status_code=423, detail="Akun terkunci sementara. Coba lagi nanti.")

    if not verify_captcha(req.captcha_id, req.captcha_answer):
        await log_security('captcha_failed', req.username, request=request)
        raise HTTPException(status_code=400, detail="Captcha salah atau kedaluwarsa")

    user = await db.users.find_one({'username': req.username})
    if not user:
        record_login_attempt(req.username, success=False)
        await log_security('login_failed', req.username, {'reason': 'user_not_found'}, request)
        raise HTTPException(status_code=401, detail="Username atau password salah")

    if not user.get('is_active', True):
        await log_security('login_failed', req.username, {'reason': 'inactive'}, request)
        raise HTTPException(status_code=403, detail="Akun Anda dinonaktifkan")

    if not verify_password(req.password, user['password_hash']):
        attempt_info = record_login_attempt(req.username, success=False)
        await log_security('login_failed', req.username,
                          {'reason': 'wrong_password', 'attempts': attempt_info.get('attempts', 0)}, request)
        if attempt_info.get('locked'):
            raise HTTPException(status_code=423, detail="Terlalu banyak percobaan gagal. Akun terkunci 15 menit.")
        raise HTTPException(status_code=401, detail=f"Username atau password salah. Sisa percobaan: {5 - attempt_info.get('attempts', 0)}")

    record_login_attempt(req.username, success=True)
    active_role = user['roles'][0] if user.get('roles') else 'guru'
    token = create_access_token({'sub': user['id'], 'username': user['username'], 'active_role': active_role})
    await db.users.update_one({'id': user['id']}, {'$set': {'last_login_at': datetime.utcnow().isoformat()}})
    await log_security('login_success', req.username, {'role': active_role}, request)
    user_clean = serialize_doc(user.copy())
    user_clean.pop('password_hash', None)
    settings = await get_settings()
    return LoginResponse(
        access_token=token, user=user_clean, active_role=active_role,
        expires_in_minutes=settings.get('session_max_hours', 12) * 60,
        idle_timeout_minutes=settings.get('idle_timeout_minutes', 30),
    )


@api_router.post("/auth/switch-role")
async def switch_role(req: RoleSwitchRequest, request: Request, user: Dict = Depends(get_current_user)):
    if req.new_role not in user.get('roles', []):
        raise HTTPException(status_code=403, detail=f"Anda tidak memiliki peran '{req.new_role}'")
    token = create_access_token({'sub': user['id'], 'username': user['username'], 'active_role': req.new_role})
    await log_audit(user, 'role_switch', 'session', details={'new_role': req.new_role}, request=request)
    user.pop('password_hash', None)
    user['active_role'] = req.new_role
    return {'access_token': token, 'token_type': 'bearer', 'active_role': req.new_role, 'user': user}


@api_router.get("/auth/me")
async def me(user: Dict = Depends(get_current_user)):
    user.pop('password_hash', None)
    return user


@api_router.post("/auth/logout")
async def logout(request: Request, user: Dict = Depends(get_current_user)):
    await log_audit(user, 'logout', 'session', request=request)
    return {'message': 'Logged out'}


# ============================================================
# SETTINGS
# ============================================================
@api_router.get("/settings")
async def public_settings():
    s = await get_settings()
    return {
        'app_name': s.get('app_name'),
        'school_name': s.get('school_name'),
        'npsn': s.get('npsn'),
        'address': s.get('address'),
        'logo_url': s.get('logo_url'),
        'favicon_url': s.get('favicon_url'),
        'primary_color': s.get('primary_color'),
        'active_days': s.get('active_days', []),
        'teaching_slots': s.get('teaching_slots', []),
        'idle_timeout_minutes': s.get('idle_timeout_minutes', 30),
        'session_max_hours': s.get('session_max_hours', 12),
    }


@api_router.get("/admin/settings")
async def get_full_settings(user: Dict = Depends(require_role('admin'))):
    return await get_settings()


@api_router.put("/admin/settings")
async def update_settings(payload: Dict[str, Any], request: Request, user: Dict = Depends(require_role('admin'))):
    payload['updated_at'] = datetime.utcnow().isoformat()
    payload['updated_by'] = user['username']
    await db.settings.update_one({'id': 'global_config'}, {'$set': payload}, upsert=True)
    await log_audit(user, 'update', 'settings', 'global_config', details={'keys': list(payload.keys())}, request=request)
    return await get_settings()


@api_router.post("/admin/settings/upload-logo")
async def upload_logo(file: UploadFile = File(...), kind: str = Form('logo'),
                      request: Request = None, user: Dict = Depends(require_role('admin'))):
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File terlalu besar (max 5MB)")
    mime = file.content_type or 'image/png'
    b64 = base64.b64encode(contents).decode('utf-8')
    data_url = f"data:{mime};base64,{b64}"
    field_map = {'logo': 'logo_url', 'favicon': 'favicon_url', 'report_logo': 'report_logo_url'}
    field = field_map.get(kind, 'logo_url')
    await db.settings.update_one({'id': 'global_config'},
                                {'$set': {field: data_url, 'updated_at': datetime.utcnow().isoformat(),
                                          'updated_by': user['username']}}, upsert=True)
    await log_audit(user, 'upload', 'settings', field, request=request)
    return {field: data_url}


# ============================================================
# ACADEMIC YEARS
# ============================================================
@api_router.get("/academic-years")
async def list_academic_years(user: Dict = Depends(get_current_user)):
    items = await db.academic_years.find({}, {'_id': 0}).sort('name', -1).to_list(100)
    return [serialize_doc(i) for i in items]


@api_router.get("/academic-years/active")
async def active_academic_year():
    ay = await get_active_academic_year()
    return ay or {}


@api_router.post("/academic-years")
async def create_academic_year(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    ay = AcademicYearModel(**payload)
    doc = ay.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if ay.is_active:
        await db.academic_years.update_many({}, {'$set': {'is_active': False}})
    await db.academic_years.insert_one(doc)
    await log_audit(user, 'create', 'academic_year', ay.id, details={'name': ay.name}, request=request)
    return serialize_doc(doc)


@api_router.put("/academic-years/{ay_id}")
async def update_academic_year(ay_id: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    """Edit TP - name, semester_type, semesters list, active_semester"""
    # Sanitize: prevent _id collision
    payload.pop('_id', None)
    payload.pop('id', None)
    # If is_active set to True here, deactivate others
    if payload.get('is_active') is True:
        await db.academic_years.update_many({'id': {'$ne': ay_id}}, {'$set': {'is_active': False}})
    res = await db.academic_years.update_one({'id': ay_id}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Tahun pelajaran tidak ditemukan")
    await log_audit(user, 'update', 'academic_year', ay_id, details={'keys': list(payload.keys())}, request=request)
    doc = await db.academic_years.find_one({'id': ay_id}, {'_id': 0})
    return serialize_doc(doc)


@api_router.put("/academic-years/{ay_id}/activate")
async def activate_academic_year(ay_id: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.academic_years.update_many({}, {'$set': {'is_active': False}})
    res = await db.academic_years.update_one({'id': ay_id}, {'$set': {'is_active': True}})
    if res.matched_count == 0:
        raise HTTPException(404, "Tahun pelajaran tidak ditemukan")
    await log_audit(user, 'activate', 'academic_year', ay_id, request=request)
    return {'message': 'Aktif', 'id': ay_id}


@api_router.delete("/academic-years/{ay_id}")
async def delete_academic_year(ay_id: str, request: Request, user: Dict = Depends(require_role('admin'))):
    res = await db.academic_years.delete_one({'id': ay_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Tidak ditemukan")
    await log_audit(user, 'delete', 'academic_year', ay_id, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# CLASSES
# ============================================================
@api_router.get("/classes")
async def list_classes(academic_year_id: Optional[str] = None, user: Dict = Depends(get_current_user)):
    q = {}
    if academic_year_id:
        q['academic_year_id'] = academic_year_id
    items = await db.classes.find(q, {'_id': 0}).sort('name', 1).to_list(500)
    return [serialize_doc(i) for i in items]


@api_router.post("/classes")
async def create_class(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    cls = ClassModel(**payload)
    doc = cls.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.classes.insert_one(doc)
    await log_audit(user, 'create', 'class', cls.id, details={'name': cls.name}, request=request)
    return serialize_doc(doc)


@api_router.put("/classes/{cid}")
async def update_class(cid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    res = await db.classes.update_one({'id': cid}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Kelas tidak ditemukan")
    await log_audit(user, 'update', 'class', cid, details=payload, request=request)
    doc = await db.classes.find_one({'id': cid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/classes/{cid}")
async def delete_class(cid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.classes.delete_one({'id': cid})
    await log_audit(user, 'delete', 'class', cid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# ROOMS
# ============================================================
@api_router.get("/rooms")
async def list_rooms(user: Dict = Depends(get_current_user)):
    items = await db.rooms.find({}, {'_id': 0}).sort('name', 1).to_list(500)
    return [serialize_doc(i) for i in items]


@api_router.post("/rooms")
async def create_room(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    if payload.get('qr_mode') == 'dynamic' and not payload.get('qr_secret'):
        payload['qr_secret'] = pyotp.random_base32()
    room = RoomModel(**payload)
    doc = room.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.rooms.insert_one(doc)
    await log_audit(user, 'create', 'room', room.id, details={'name': room.name}, request=request)
    return serialize_doc(doc)


@api_router.put("/rooms/{rid}")
async def update_room(rid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    if payload.get('qr_mode') == 'dynamic' and not payload.get('qr_secret'):
        existing = await db.rooms.find_one({'id': rid})
        if existing and not existing.get('qr_secret'):
            payload['qr_secret'] = pyotp.random_base32()
    res = await db.rooms.update_one({'id': rid}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Ruangan tidak ditemukan")
    await log_audit(user, 'update', 'room', rid, details=payload, request=request)
    doc = await db.rooms.find_one({'id': rid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/rooms/{rid}")
async def delete_room(rid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.rooms.delete_one({'id': rid})
    await log_audit(user, 'delete', 'room', rid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# SUBJECTS
# ============================================================
@api_router.get("/subjects")
async def list_subjects(user: Dict = Depends(get_current_user)):
    items = await db.subjects.find({}, {'_id': 0}).sort('name', 1).to_list(500)
    return [serialize_doc(i) for i in items]


@api_router.post("/subjects")
async def create_subject(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    sub = SubjectModel(**payload)
    doc = sub.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.subjects.insert_one(doc)
    await log_audit(user, 'create', 'subject', sub.id, details={'name': sub.name}, request=request)
    return serialize_doc(doc)


@api_router.put("/subjects/{sid}")
async def update_subject(sid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    res = await db.subjects.update_one({'id': sid}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Mata pelajaran tidak ditemukan")
    await log_audit(user, 'update', 'subject', sid, details=payload, request=request)
    doc = await db.subjects.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/subjects/{sid}")
async def delete_subject(sid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.subjects.delete_one({'id': sid})
    await log_audit(user, 'delete', 'subject', sid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# USERS
# ============================================================
@api_router.get("/users")
async def list_users(role: Optional[str] = None, user: Dict = Depends(require_role('admin'))):
    q = {}
    if role:
        q['roles'] = role
    items = await db.users.find(q, {'_id': 0, 'password_hash': 0}).to_list(2000)
    return [serialize_doc(i) for i in items]


@api_router.post("/users")
async def create_user(req: UserCreateRequest, request: Request, user: Dict = Depends(require_role('admin'))):
    existing = await db.users.find_one({'username': req.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username sudah digunakan")
    invalid = [r for r in req.roles if r not in ROLES]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Peran tidak valid: {invalid}")
    u = UserModel(
        username=req.username, password_hash=hash_password(req.password),
        full_name=req.full_name, nip_nuptk=req.nip_nuptk, nisn=req.nisn,
        email=req.email, phone=req.phone, roles=req.roles,
        homeroom_class_id=req.homeroom_class_id, student_class_id=req.student_class_id,
        parent_of=req.parent_of,
    )
    doc = u.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    await log_audit(user, 'create', 'user', u.id, details={'username': req.username, 'roles': req.roles}, request=request)
    doc.pop('password_hash', None)
    return serialize_doc(doc)


@api_router.put("/users/{uid}")
async def update_user(uid: str, req: UserUpdateRequest, request: Request, user: Dict = Depends(require_role('admin'))):
    update = {k: v for k, v in req.model_dump(exclude_unset=True).items() if v is not None}
    if 'new_password' in update:
        update['password_hash'] = hash_password(update.pop('new_password'))
    res = await db.users.update_one({'id': uid}, {'$set': update})
    if res.matched_count == 0:
        raise HTTPException(404, "User tidak ditemukan")
    await log_audit(user, 'update', 'user', uid, details={'keys': list(update.keys())}, request=request)
    doc = await db.users.find_one({'id': uid}, {'_id': 0, 'password_hash': 0})
    return serialize_doc(doc)


@api_router.delete("/users/{uid}")
async def delete_user(uid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.users.delete_one({'id': uid})
    await log_audit(user, 'delete', 'user', uid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# SCHEDULES
# ============================================================
@api_router.get("/schedules")
async def list_schedules(
    academic_year_id: Optional[str] = None, semester: Optional[str] = None,
    class_id: Optional[str] = None, teacher_id: Optional[str] = None, day: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    q = {}
    if academic_year_id: q['academic_year_id'] = academic_year_id
    if semester: q['semester'] = semester
    if class_id: q['class_id'] = class_id
    if teacher_id: q['teacher_id'] = teacher_id
    if day: q['day'] = day
    items = await db.schedules.find(q, {'_id': 0}).sort([('day', 1), ('start_time', 1)]).to_list(2000)
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['subject_code'] = sub.get('code') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        enriched.append(serialize_doc(s))
    return enriched


@api_router.post("/schedules")
async def create_schedule(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    sched = ScheduleModel(**payload)
    doc = sched.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.schedules.insert_one(doc)
    await log_audit(user, 'create', 'schedule', sched.id, request=request)
    return serialize_doc(doc)


@api_router.put("/schedules/{sid}")
async def update_schedule(sid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    res = await db.schedules.update_one({'id': sid}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Jadwal tidak ditemukan")
    await log_audit(user, 'update', 'schedule', sid, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/schedules/{sid}")
async def delete_schedule(sid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.schedules.delete_one({'id': sid})
    await log_audit(user, 'delete', 'schedule', sid, request=request)
    return {'message': 'Dihapus'}


@api_router.get("/schedules/my-today")
async def my_today_schedule(user: Dict = Depends(get_current_user)):
    ay = await get_active_academic_year()
    if not ay:
        return []
    day = current_day_id()
    items = await db.schedules.find({
        'teacher_id': user['id'], 'day': day, 'academic_year_id': ay['id'],
    }, {'_id': 0}).sort('start_time', 1).to_list(50)
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        journal = await db.journals.find_one({'schedule_id': s['id'], 'teacher_id': user['id']}, {'_id': 0, 'id': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['room_name'] = room.get('name') if room else None
        s['journal_filled'] = bool(journal)
        s['journal_id'] = journal.get('id') if journal else None
        enriched.append(serialize_doc(s))
    return enriched


# ============================================================
# QR CODE
# ============================================================
@api_router.get("/rooms/{rid}/qr")
async def get_room_qr(rid: str, mode: str = 'static', user: Dict = Depends(get_current_user)):
    room = await db.rooms.find_one({'id': rid}, {'_id': 0})
    if not room:
        raise HTTPException(404, "Ruangan tidak ditemukan")
    if mode == 'dynamic':
        if not room.get('qr_secret'):
            new_secret = pyotp.random_base32()
            await db.rooms.update_one({'id': rid}, {'$set': {'qr_secret': new_secret}})
            room['qr_secret'] = new_secret
        token = generate_dynamic_qr_payload(rid, room['qr_secret'])
    else:
        token = encrypt_qr_payload(rid)
    qr_b64 = generate_qr_image_b64(token, size=10)
    return {
        'token': token, 'qr_image_b64': qr_b64, 'mode': mode,
        'room_id': rid, 'room_name': room.get('name'),
        'refresh_seconds': 30 if mode == 'dynamic' else 0,
    }


@api_router.post("/rooms/{rid}/qr-card")
async def generate_qr_card(rid: str, template_id: Optional[str] = Form(None),
                            class_name: Optional[str] = Form(None),
                            user: Dict = Depends(require_role('admin'))):
    room = await db.rooms.find_one({'id': rid}, {'_id': 0})
    if not room:
        raise HTTPException(404, "Ruangan tidak ditemukan")
    settings = await get_settings()
    token = encrypt_qr_payload(rid)
    template_bytes = None
    if template_id:
        tpl = await db.qr_templates.find_one({'id': template_id}, {'_id': 0})
        if tpl and tpl.get('image_b64'):
            b64 = tpl['image_b64']
            if ',' in b64:
                b64 = b64.split(',', 1)[1]
            template_bytes = base64.b64decode(b64)
    if not class_name:
        cls = await db.classes.find_one({'room_id': rid}, {'_id': 0})
        class_name = cls.get('name') if cls else room.get('name', '')
    png_bytes = create_b5_card(
        qr_data=token, room_name=room.get('name', rid), class_name=class_name,
        template_bytes=template_bytes,
        school_name=settings.get('school_name', 'MTsN 2 Kota Malang'),
        app_name=settings.get('app_name', 'Super Apps MATSANDATAMA'),
    )
    return StreamingResponse(io.BytesIO(png_bytes), media_type="image/png",
                             headers={"Content-Disposition": f'inline; filename="qr-card-{rid}.png"'})


@api_router.get("/qr-templates")
async def list_qr_templates(user: Dict = Depends(require_role('admin'))):
    items = await db.qr_templates.find({}, {'_id': 0}).to_list(100)
    return [serialize_doc(i) for i in items]


@api_router.post("/qr-templates")
async def upload_qr_template(file: UploadFile = File(...), name: str = Form(...),
                              request: Request = None, user: Dict = Depends(require_role('admin'))):
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File terlalu besar (max 10MB)")
    b64 = base64.b64encode(contents).decode('utf-8')
    mime = file.content_type or 'image/png'
    tpl = QRTemplateModel(name=name, image_b64=f"data:{mime};base64,{b64}")
    doc = tpl.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.qr_templates.insert_one(doc)
    await log_audit(user, 'upload', 'qr_template', tpl.id, details={'name': name}, request=request)
    return serialize_doc(doc)


@api_router.delete("/qr-templates/{tid}")
async def delete_qr_template(tid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.qr_templates.delete_one({'id': tid})
    await log_audit(user, 'delete', 'qr_template', tid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# JURNAL
# ============================================================
async def _validate_qr_full(qr_token: str, user_lat: Optional[float], user_lon: Optional[float],
                              teacher_id: str) -> Dict:
    settings = await get_settings()
    result = {
        'overall_valid': False,
        'qr': {'valid': False, 'reason': 'Belum diperiksa'},
        'schedule': {'valid': False, 'reason': 'Belum diperiksa'},
        'gps': {'valid': False, 'reason': 'Belum diperiksa'},
        'context': {},
    }
    payload = decrypt_qr_payload(qr_token)
    if not payload:
        result['qr'] = {'valid': False, 'reason': 'QR Code tidak valid atau kedaluwarsa'}
        return result
    room_id = payload.get('room_id')
    result['qr'] = {'valid': True, 'reason': 'QR berhasil divalidasi', 'mode': payload.get('mode', 'static')}

    room = await db.rooms.find_one({'id': room_id}, {'_id': 0})
    if not room:
        result['qr'] = {'valid': False, 'reason': 'Ruangan tidak terdaftar dalam sistem'}
        return result

    if payload.get('mode') == 'dynamic':
        secret = room.get('qr_secret')
        if not secret:
            result['qr'] = {'valid': False, 'reason': 'Ruangan tidak memiliki secret untuk QR dinamis'}
            return result
        dyn = validate_dynamic_qr(qr_token, secret)
        if not dyn['valid']:
            result['qr'] = dyn
            return result

    ay = await get_active_academic_year()
    if not ay:
        result['schedule'] = {'valid': False, 'reason': 'Tidak ada tahun pelajaran aktif'}
        return result
    schedules = await db.schedules.find({
        'academic_year_id': ay['id'], 'teacher_id': teacher_id, 'room_id': room_id,
    }, {'_id': 0}).to_list(100)
    for s in schedules:
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1})
        if sub:
            s['subject_name'] = sub.get('name')
    grace = settings.get('grace_minutes', 15)
    sched_result = validate_schedule(teacher_id, room_id, schedules, grace_minutes=grace)
    result['schedule'] = sched_result
    if not sched_result['valid']:
        return result

    gps_enabled = room.get('gps_enabled', settings.get('gps_default_enabled', True))
    radius = room.get('gps_radius_meters', settings.get('gps_default_radius', 20))
    gps_result = validate_gps(user_lat, user_lon, room.get('gps_lat'), room.get('gps_lon'),
                              radius, gps_enabled=gps_enabled)
    result['gps'] = gps_result
    if not gps_result['valid']:
        return result

    result['overall_valid'] = True
    result['context'] = {
        'room': serialize_doc(room),
        'schedule': sched_result.get('schedule'),
        'start_time': sched_result.get('start_time'),
        'end_time': sched_result.get('end_time'),
    }
    return result


@api_router.post("/jurnal/validate")
async def validate_qr(req: QRValidateRequest, user: Dict = Depends(get_current_user)):
    return await _validate_qr_full(req.qr_token, req.user_lat, req.user_lon, user['id'])


@api_router.post("/jurnal")
async def create_journal(req: JournalCreateRequest, request: Request, user: Dict = Depends(get_current_user)):
    validation = await _validate_qr_full(req.qr_token, req.user_lat, req.user_lon, user['id'])
    if not validation['overall_valid']:
        await log_security('journal_blocked', user['username'],
                          {'reasons': [v.get('reason') for k, v in validation.items() if isinstance(v, dict) and not v.get('valid', True)]},
                          request)
        raise HTTPException(status_code=400, detail={'message': 'Validasi gagal', 'validation': validation})

    sched = validation['context']['schedule']
    room = validation['context']['room']
    ay = await get_active_academic_year()

    existing = await db.journals.find_one({'schedule_id': sched['id'], 'teacher_id': user['id']})
    if existing:
        raise HTTPException(status_code=400, detail="Jurnal untuk jadwal ini sudah diisi")

    journal = JournalModel(
        schedule_id=sched['id'], teacher_id=user['id'], class_id=sched['class_id'],
        subject_id=sched['subject_id'], room_id=room['id'], academic_year_id=ay['id'],
        semester=sched.get('semester', 'ganjil'),
        materi=req.materi, catatan=req.catatan,
        siswa_hadir=req.siswa_hadir, siswa_tidak_hadir=req.siswa_tidak_hadir,
        siswa_izin=req.siswa_izin, siswa_sakit=req.siswa_sakit,
        scheduled_start=validation['context'].get('start_time'),
        scheduled_end=validation['context'].get('end_time'),
        validations=validation,
        qr_mode=validation['qr'].get('mode', 'static'),
    )
    doc = journal.model_dump()
    if isinstance(doc.get('started_at'), datetime):
        doc['started_at'] = doc['started_at'].isoformat()
    if isinstance(doc.get('created_at'), datetime):
        doc['created_at'] = doc['created_at'].isoformat()
    await db.journals.insert_one(doc)
    await log_audit(user, 'create', 'journal', journal.id,
                   details={'class_id': sched['class_id'], 'subject_id': sched['subject_id']}, request=request)
    return serialize_doc(doc)


@api_router.get("/jurnal/my")
async def my_journals(user: Dict = Depends(get_current_user)):
    items = await db.journals.find({'teacher_id': user['id']}, {'_id': 0}).sort('started_at', -1).to_list(500)
    enriched = []
    for j in items:
        cls = await db.classes.find_one({'id': j.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': j.get('subject_id')}, {'_id': 0, 'name': 1})
        room = await db.rooms.find_one({'id': j.get('room_id')}, {'_id': 0, 'name': 1})
        j['class_name'] = cls.get('name') if cls else None
        j['subject_name'] = sub.get('name') if sub else None
        j['room_name'] = room.get('name') if room else None
        enriched.append(serialize_doc(j))
    return enriched


@api_router.get("/jurnal/by-class/{class_id}")
async def journals_by_class(class_id: str, limit: int = 50, user: Dict = Depends(get_current_user)):
    items = await db.journals.find({'class_id': class_id}, {'_id': 0}).sort('started_at', -1).to_list(limit)
    enriched = []
    for j in items:
        sub = await db.subjects.find_one({'id': j.get('subject_id')}, {'_id': 0, 'name': 1})
        teacher = await db.users.find_one({'id': j.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        j['subject_name'] = sub.get('name') if sub else None
        j['teacher_name'] = teacher.get('full_name') if teacher else None
        enriched.append(serialize_doc(j))
    return enriched


# ============================================================
# PUBLIC MONITORING
# ============================================================
@api_router.get("/public/monitoring")
async def public_monitoring(day: Optional[str] = None):
    ay = await db.academic_years.find_one({'is_active': True}, {'_id': 0})
    if not ay:
        return {'time': now_wib().isoformat(), 'day': current_day_id(), 'classes': [], 'active_year': None,
                'stats': {'total': 0, 'filled': 0, 'pending': 0, 'missing': 0, 'upcoming': 0}}

    current_day = day or current_day_id()
    schedules = await db.schedules.find({
        'academic_year_id': ay['id'], 'day': current_day,
    }, {'_id': 0}).sort('start_time', 1).to_list(2000)

    now = now_wib()
    items = []
    for s in schedules:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        journal = await db.journals.find_one({'schedule_id': s['id']}, {'_id': 0, 'id': 1, 'materi': 1, 'started_at': 1})
        try:
            sh, sm = map(int, s['start_time'].split(':'))
            eh, em = map(int, s['end_time'].split(':'))
            start_dt = now.replace(hour=sh, minute=sm, second=0, microsecond=0)
            end_dt = now.replace(hour=eh, minute=em, second=0, microsecond=0)
        except Exception:
            continue
        if now < start_dt:
            status_label = 'upcoming'
        elif start_dt <= now <= end_dt:
            status_label = 'active'
        else:
            status_label = 'past'
        if journal:
            jurnal_status = 'filled'
        elif status_label == 'past':
            jurnal_status = 'missing'
        elif status_label == 'active':
            jurnal_status = 'pending'
        else:
            jurnal_status = 'not_started'
        items.append({
            'schedule_id': s['id'],
            'class_name': cls.get('name') if cls else '-',
            'subject_name': sub.get('name') if sub else '-',
            'teacher_name': teacher.get('full_name') if teacher else '-',
            'room_name': room.get('name') if room else '-',
            'start_time': s['start_time'], 'end_time': s['end_time'],
            'status': status_label, 'jurnal_status': jurnal_status,
            'jurnal_materi': journal.get('materi') if journal else None,
            'jurnal_filled_at': journal.get('started_at') if journal else None,
        })

    settings = await get_settings()
    return {
        'time': now.isoformat(), 'day': current_day,
        'academic_year': ay.get('name'),
        'school_name': settings.get('school_name'),
        'app_name': settings.get('app_name'),
        'logo_url': settings.get('logo_url'),
        'classes': items,
        'stats': {
            'total': len(items),
            'filled': len([i for i in items if i['jurnal_status'] == 'filled']),
            'pending': len([i for i in items if i['jurnal_status'] == 'pending']),
            'missing': len([i for i in items if i['jurnal_status'] == 'missing']),
            'upcoming': len([i for i in items if i['jurnal_status'] == 'not_started']),
        }
    }


# ============================================================
# WALI KELAS
# ============================================================
@api_router.get("/wali-kelas/my-class")
async def wali_kelas_dashboard(user: Dict = Depends(get_current_user)):
    cls = await db.classes.find_one({'homeroom_teacher_id': user['id']}, {'_id': 0})
    if not cls:
        return {'class': None, 'today_schedule': [], 'students': []}
    day = current_day_id()
    ay = await get_active_academic_year()
    schedules = await db.schedules.find({
        'class_id': cls['id'], 'day': day,
        'academic_year_id': ay['id'] if ay else None,
    }, {'_id': 0}).sort('start_time', 1).to_list(50)
    enriched = []
    for s in schedules:
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        journal = await db.journals.find_one({'schedule_id': s['id']}, {'_id': 0, 'id': 1, 'materi': 1})
        s['subject_name'] = sub.get('name') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        s['journal_filled'] = bool(journal)
        s['journal_materi'] = journal.get('materi') if journal else None
        enriched.append(serialize_doc(s))
    students = await db.users.find({'student_class_id': cls['id'], 'roles': 'siswa'},
                                    {'_id': 0, 'password_hash': 0}).to_list(200)
    return {
        'class': serialize_doc(cls),
        'today_schedule': enriched,
        'students': [serialize_doc(s) for s in students],
    }


# ============================================================
# PARENT/STUDENT
# ============================================================
@api_router.get("/parent/children")
async def parent_children(user: Dict = Depends(get_current_user)):
    if 'orang_tua' not in user.get('roles', []):
        return []
    ids = user.get('parent_of', [])
    if not ids:
        return []
    items = await db.users.find({'id': {'$in': ids}}, {'_id': 0, 'password_hash': 0}).to_list(20)
    return [serialize_doc(i) for i in items]


@api_router.get("/student/{student_id}/today")
async def student_today(student_id: str, user: Dict = Depends(get_current_user)):
    student = await db.users.find_one({'id': student_id}, {'_id': 0, 'password_hash': 0})
    if not student:
        raise HTTPException(404, "Siswa tidak ditemukan")
    is_self = user['id'] == student_id
    is_parent = student_id in user.get('parent_of', [])
    is_admin = 'admin' in user.get('roles', [])
    cls = await db.classes.find_one({'id': student.get('student_class_id')}, {'_id': 0})
    is_homeroom = cls and cls.get('homeroom_teacher_id') == user['id']
    if not (is_self or is_parent or is_admin or is_homeroom):
        raise HTTPException(403, "Tidak diizinkan melihat data siswa ini")
    if not cls:
        return {'student': serialize_doc(student), 'class': None, 'today_schedule': []}
    day = current_day_id()
    ay = await get_active_academic_year()
    schedules = await db.schedules.find({
        'class_id': cls['id'], 'day': day,
        'academic_year_id': ay['id'] if ay else None,
    }, {'_id': 0}).sort('start_time', 1).to_list(50)
    enriched = []
    for s in schedules:
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        journal = await db.journals.find_one({'schedule_id': s['id']}, {'_id': 0, 'materi': 1, 'catatan': 1, 'started_at': 1})
        s['subject_name'] = sub.get('name') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        s['journal'] = serialize_doc(journal) if journal else None
        enriched.append(serialize_doc(s))
    return {'student': serialize_doc(student), 'class': serialize_doc(cls), 'today_schedule': enriched}


# ============================================================
# LOGS / STATS
# ============================================================
@api_router.get("/admin/audit-logs")
async def get_audit_logs(limit: int = 200, user: Dict = Depends(require_role('admin'))):
    items = await db.audit_logs.find({}, {'_id': 0}).sort('timestamp', -1).to_list(limit)
    return [serialize_doc(i) for i in items]


@api_router.get("/admin/security-logs")
async def get_security_logs(limit: int = 200, user: Dict = Depends(require_role('admin'))):
    items = await db.security_logs.find({}, {'_id': 0}).sort('timestamp', -1).to_list(limit)
    return [serialize_doc(i) for i in items]


@api_router.get("/roles")
async def get_roles():
    return [{'value': r, 'label': ROLE_LABELS[r]} for r in ROLES]


@api_router.get("/admin/stats")
async def admin_stats(user: Dict = Depends(require_role('admin'))):
    today = current_day_id()
    ay = await get_active_academic_year()
    total_users = await db.users.count_documents({})
    total_classes = await db.classes.count_documents({})
    total_rooms = await db.rooms.count_documents({})
    total_schedules_today = await db.schedules.count_documents({
        'day': today, 'academic_year_id': ay['id'] if ay else 'none',
    })
    now = now_wib()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    total_journals_today = await db.journals.count_documents({'started_at': {'$gte': today_start}})
    return {
        'total_users': total_users, 'total_classes': total_classes, 'total_rooms': total_rooms,
        'total_schedules_today': total_schedules_today, 'total_journals_today': total_journals_today,
        'active_academic_year': ay.get('name') if ay else None, 'current_day': today,
    }


# ============================================================
# STUDENT DATA (Data Siswa)
# ============================================================
async def _user_can_view_class(user: Dict, class_id: str) -> bool:
    if 'admin' in user.get('roles', []):
        return True
    cls = await db.classes.find_one({'id': class_id}, {'_id': 0})
    if cls and cls.get('homeroom_teacher_id') == user['id']:
        return True
    # Allow guru_bk, guru_tata_tertib, guru_piket to view all
    overlap = set(user.get('roles', [])) & {'guru_bk', 'guru_tata_tertib', 'guru_piket', 'tenaga_kependidikan'}
    if overlap:
        return True
    return False


@api_router.get("/students")
async def list_students(class_id: Optional[str] = None, user: Dict = Depends(get_current_user)):
    """Get list of students. Admin sees all; wali kelas sees own class; siswa sees self only."""
    if 'siswa' in user.get('roles', []) and len(user.get('roles', [])) == 1:
        # Pure student can only see themselves
        me = await db.users.find_one({'id': user['id']}, {'_id': 0, 'password_hash': 0})
        return [serialize_doc(me)] if me else []
    q = {'roles': 'siswa'}
    if class_id:
        if not await _user_can_view_class(user, class_id):
            raise HTTPException(403, "Tidak diizinkan melihat siswa kelas ini")
        q['student_class_id'] = class_id
    elif 'admin' not in user.get('roles', []):
        # non-admin must specify a class they have access to
        cls = await db.classes.find_one({'homeroom_teacher_id': user['id']}, {'_id': 0, 'id': 1})
        if cls:
            q['student_class_id'] = cls['id']
        else:
            return []
    items = await db.users.find(q, {'_id': 0, 'password_hash': 0}).sort('full_name', 1).to_list(2000)
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('student_class_id')}, {'_id': 0, 'name': 1})
        s['class_name'] = cls.get('name') if cls else None
        enriched.append(serialize_doc(s))
    return enriched


# ============================================================
# CLASS ATTENDANCE (Kehadiran Siswa)
# ============================================================
@api_router.get("/attendance/class/{class_id}")
async def get_class_attendance(class_id: str, date: Optional[str] = None,
                                user: Dict = Depends(get_current_user)):
    if not await _user_can_view_class(user, class_id):
        raise HTTPException(403, "Tidak diizinkan")
    q = {'class_id': class_id}
    if date:
        q['date'] = date
    items = await db.class_attendance.find(q, {'_id': 0}).sort('date', -1).to_list(200)
    return [serialize_doc(i) for i in items]


@api_router.post("/attendance/class")
async def submit_class_attendance(req: ClassAttendanceSubmit, request: Request,
                                   user: Dict = Depends(get_current_user)):
    if not await _user_can_view_class(user, req.class_id):
        raise HTTPException(403, "Tidak diizinkan")
    # Calculate summary
    summary = {'hadir': 0, 'sakit': 0, 'izin': 0, 'alpa': 0}
    for r in req.records:
        st = r.get('status', 'hadir')
        summary[st] = summary.get(st, 0) + 1
    # Upsert by class_id + date
    existing = await db.class_attendance.find_one({'class_id': req.class_id, 'date': req.date})
    doc = {
        'class_id': req.class_id, 'date': req.date,
        'records': req.records, 'recorded_by': user['id'],
        'recorded_at': datetime.utcnow().isoformat(), 'summary': summary,
    }
    if existing:
        await db.class_attendance.update_one({'_id': existing['_id']}, {'$set': doc})
        doc['id'] = existing.get('id', str(uuid.uuid4()))
    else:
        doc['id'] = str(uuid.uuid4())
        await db.class_attendance.insert_one(doc)
    await log_audit(user, 'submit', 'class_attendance', doc['id'],
                   details={'class_id': req.class_id, 'date': req.date, 'summary': summary},
                   request=request)
    return serialize_doc(doc)


# ============================================================
# CLASS CLEANLINESS (Kebersihan Kelas)
# ============================================================
@api_router.get("/cleanliness/class/{class_id}")
async def get_class_cleanliness(class_id: str, limit: int = 30,
                                 user: Dict = Depends(get_current_user)):
    if not await _user_can_view_class(user, class_id):
        raise HTTPException(403, "Tidak diizinkan")
    items = await db.class_cleanliness.find({'class_id': class_id}, {'_id': 0}).sort('date', -1).to_list(limit)
    return [serialize_doc(i) for i in items]


@api_router.post("/cleanliness/class")
async def submit_class_cleanliness(req: ClassCleanlinessSubmit, request: Request,
                                    user: Dict = Depends(get_current_user)):
    if not await _user_can_view_class(user, req.class_id):
        raise HTTPException(403, "Tidak diizinkan")
    existing = await db.class_cleanliness.find_one({'class_id': req.class_id, 'date': req.date})
    doc = req.model_dump()
    doc['recorded_by'] = user['id']
    doc['recorded_at'] = datetime.utcnow().isoformat()
    if existing:
        await db.class_cleanliness.update_one({'_id': existing['_id']}, {'$set': doc})
        doc['id'] = existing.get('id', str(uuid.uuid4()))
    else:
        doc['id'] = str(uuid.uuid4())
        await db.class_cleanliness.insert_one(doc)
    await log_audit(user, 'submit', 'class_cleanliness', doc['id'],
                   details={'class_id': req.class_id, 'date': req.date, 'condition': req.condition},
                   request=request)
    return serialize_doc(doc)


# ============================================================
# SCHEDULE EXCEL IMPORT
# ============================================================
@api_router.get("/schedules/excel-template")
async def schedule_excel_template(user: Dict = Depends(require_role('admin'))):
    """Download Excel template for bulk schedule import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal"
    headers = ['hari', 'jam_mulai', 'jam_selesai', 'kelas', 'mapel_kode', 'guru_username', 'ruang_kode', 'semester']
    ws.append(headers)
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='006837')
        cell.alignment = Alignment(horizontal='center')
    # Example rows
    examples = [
        ['senin', '07:00', '07:45', '7A', 'MTK', 'guru1', 'R-7A', 'ganjil'],
        ['senin', '07:45', '08:30', '7A', 'IPA', 'walas7a', 'R-7A', 'ganjil'],
        ['selasa', '07:00', '07:45', '7B', 'MTK', 'guru1', 'R-7B', 'ganjil'],
    ]
    for row in examples:
        ws.append(row)
    for col_letter, width in zip('ABCDEFGH', [10, 12, 12, 10, 12, 18, 12, 12]):
        ws.column_dimensions[col_letter].width = width

    # Add INSTRUKSI sheet
    ws2 = wb.create_sheet("INSTRUKSI")
    ws2.append(["Petunjuk Pengisian Template Jadwal"])
    ws2['A1'].font = Font(bold=True, size=14)
    instructions = [
        "",
        "1. Isi data jadwal pada sheet 'Jadwal' mulai baris 2.",
        "2. Kolom 'hari' wajib salah satu dari: senin, selasa, rabu, kamis, jumat, sabtu (huruf kecil).",
        "3. Kolom 'jam_mulai' dan 'jam_selesai' format HH:MM (24-jam), contoh: 07:00, 13:45.",
        "4. Kolom 'kelas' diisi NAMA kelas seperti yang terdaftar (contoh: 7A, 8B).",
        "5. Kolom 'mapel_kode' diisi KODE mapel (contoh: MTK, IPA, BIN).",
        "6. Kolom 'guru_username' diisi USERNAME guru pengampu.",
        "7. Kolom 'ruang_kode' diisi NAMA ruangan (contoh: R-7A).",
        "8. Kolom 'semester' diisi 'ganjil' atau 'genap' (untuk regular), atau '1','2','3','4','5','6' (untuk percepatan).",
        "9. Sistem akan mengabaikan baris kosong dan menolak baris dengan data tidak ditemukan.",
        "10. Setelah selesai, upload file melalui menu Admin > Jadwal > Import Excel.",
    ]
    for line in instructions:
        ws2.append([line])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="template_jadwal_matsandatama.xlsx"'},
    )


@api_router.post("/schedules/import-excel")
async def schedule_import_excel(file: UploadFile = File(...), request: Request = None,
                                 user: Dict = Depends(require_role('admin'))):
    """Bulk import schedules from Excel file"""
    from openpyxl import load_workbook
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya file .xlsx yang didukung")
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(400, "File terlalu besar (max 5MB)")
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb['Jadwal'] if 'Jadwal' in wb.sheetnames else wb.active
    except Exception as e:
        raise HTTPException(400, f"Gagal membaca Excel: {e}")

    ay = await get_active_academic_year()
    if not ay:
        raise HTTPException(400, "Tidak ada tahun pelajaran aktif")
    classes_map = {c['name']: c['id'] for c in await db.classes.find({'academic_year_id': ay['id']}, {'_id': 0}).to_list(500)}
    subjects_map = {s['code'].upper(): s['id'] for s in await db.subjects.find({}, {'_id': 0}).to_list(500)}
    rooms_map = {r['name']: r['id'] for r in await db.rooms.find({}, {'_id': 0}).to_list(500)}
    users_map = {u['username']: u['id'] for u in await db.users.find({}, {'_id': 0, 'username': 1, 'id': 1}).to_list(2000)}
    valid_days = {'senin', 'selasa', 'rabu', 'kamis', 'jumat', 'sabtu'}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success = 0
    errors = []
    new_docs = []
    for idx, row in enumerate(rows, start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        try:
            hari = str(row[0]).strip().lower() if row[0] else ''
            jm = str(row[1]).strip() if row[1] else ''
            js = str(row[2]).strip() if row[2] else ''
            kls = str(row[3]).strip() if row[3] else ''
            mp_kode = str(row[4]).strip().upper() if row[4] else ''
            gr_username = str(row[5]).strip() if row[5] else ''
            ruang = str(row[6]).strip() if row[6] else ''
            sem = str(row[7]).strip().lower() if row[7] else 'ganjil'

            # Handle datetime objects from Excel (HH:MM:SS)
            if ':' not in jm:
                jm = str(jm)
            if ':' not in js:
                js = str(js)
            jm = jm[:5]  # take HH:MM
            js = js[:5]

            if hari not in valid_days:
                errors.append(f"Baris {idx}: hari '{hari}' tidak valid")
                continue
            if kls not in classes_map:
                errors.append(f"Baris {idx}: kelas '{kls}' tidak ditemukan")
                continue
            if mp_kode not in subjects_map:
                errors.append(f"Baris {idx}: mapel '{mp_kode}' tidak ditemukan")
                continue
            if gr_username not in users_map:
                errors.append(f"Baris {idx}: guru '{gr_username}' tidak ditemukan")
                continue
            if ruang not in rooms_map:
                errors.append(f"Baris {idx}: ruang '{ruang}' tidak ditemukan")
                continue

            sched = {
                'id': str(uuid.uuid4()),
                'academic_year_id': ay['id'], 'semester': sem,
                'class_id': classes_map[kls], 'subject_id': subjects_map[mp_kode],
                'teacher_id': users_map[gr_username], 'room_id': rooms_map[ruang],
                'day': hari, 'start_time': jm, 'end_time': js,
                'is_published': True, 'created_at': datetime.utcnow().isoformat(),
            }
            new_docs.append(sched)
            success += 1
        except Exception as e:
            errors.append(f"Baris {idx}: {e}")
    if new_docs:
        await db.schedules.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'schedule', None,
                   details={'success': success, 'errors': len(errors), 'filename': file.filename},
                   request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


# ============================================================
# SCHEDULE GRID (per hari aktif + jam slot)
# ============================================================
@api_router.get("/schedules/grid")
async def schedules_grid(class_id: Optional[str] = None, teacher_id: Optional[str] = None,
                          user: Dict = Depends(get_current_user)):
    """Return schedule data structured as grid: days x slots"""
    settings = await get_settings()
    ay = await get_active_academic_year()
    if not ay:
        return {'days': [], 'slots': [], 'grid': {}}
    active_days = settings.get('active_days', ['senin','selasa','rabu','kamis','jumat'])
    slots = settings.get('teaching_slots', [])

    q = {'academic_year_id': ay['id']}
    if class_id: q['class_id'] = class_id
    if teacher_id: q['teacher_id'] = teacher_id
    items = await db.schedules.find(q, {'_id': 0}).to_list(2000)
    # Enrich
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['subject_code'] = sub.get('code') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        enriched.append(serialize_doc(s))
    # Build grid: {day: {start_time: schedule}}
    grid = {d: {} for d in active_days}
    for s in enriched:
        d = s['day']
        if d in grid:
            grid[d][s['start_time']] = s
    return {'days': active_days, 'slots': slots, 'grid': grid, 'schedules': enriched}


# ============================================================
# JADWAL PIKET (Duty Schedule)
# ============================================================
@api_router.get("/piket-schedules")
async def list_piket(day: Optional[str] = None, user: Dict = Depends(get_current_user)):
    q = {}
    if day:
        q['day'] = day
    items = await db.piket_schedules.find(q, {'_id': 0}).sort([('day', 1), ('start_time', 1)]).to_list(500)
    enriched = []
    for s in items:
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        enriched.append(serialize_doc(s))
    return enriched


@api_router.post("/piket-schedules")
async def create_piket(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    doc = {
        'id': str(uuid.uuid4()),
        'day': payload.get('day', 'senin'),
        'shift': payload.get('shift', 'pagi'),  # pagi/siang/sore
        'start_time': payload.get('start_time', '06:30'),
        'end_time': payload.get('end_time', '14:00'),
        'teacher_id': payload.get('teacher_id'),
        'notes': payload.get('notes', ''),
        'is_active': payload.get('is_active', True),
        'created_at': datetime.utcnow().isoformat(),
    }
    await db.piket_schedules.insert_one(doc)
    await log_audit(user, 'create', 'piket_schedule', doc['id'],
                   details={'day': doc['day'], 'teacher_id': doc['teacher_id']}, request=request)
    return serialize_doc(doc)


@api_router.put("/piket-schedules/{pid}")
async def update_piket(pid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    payload.pop('_id', None)
    payload.pop('id', None)
    res = await db.piket_schedules.update_one({'id': pid}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Jadwal piket tidak ditemukan")
    await log_audit(user, 'update', 'piket_schedule', pid, details=payload, request=request)
    doc = await db.piket_schedules.find_one({'id': pid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/piket-schedules/{pid}")
async def delete_piket(pid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.piket_schedules.delete_one({'id': pid})
    await log_audit(user, 'delete', 'piket_schedule', pid, request=request)
    return {'message': 'Dihapus'}


@api_router.get("/piket-schedules/today")
async def piket_today(user: Dict = Depends(get_current_user)):
    day = current_day_id()
    items = await db.piket_schedules.find({'day': day, 'is_active': True}, {'_id': 0}).sort('start_time', 1).to_list(50)
    enriched = []
    for s in items:
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        enriched.append(serialize_doc(s))
    return enriched


# ============================================================
# ADMIN JURNAL REKAP (Data Jurnal lengkap)
# ============================================================
@api_router.get("/admin/jurnal")
async def admin_jurnal_rekap(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    class_id: Optional[str] = None,
    teacher_id: Optional[str] = None,
    subject_id: Optional[str] = None,
    academic_year_id: Optional[str] = None,
    limit: int = 500,
    user: Dict = Depends(require_role('admin'))
):
    """Rekap lengkap data jurnal mengajar untuk admin"""
    q = {}
    if class_id: q['class_id'] = class_id
    if teacher_id: q['teacher_id'] = teacher_id
    if subject_id: q['subject_id'] = subject_id
    if academic_year_id: q['academic_year_id'] = academic_year_id
    if start_date or end_date:
        date_q = {}
        if start_date: date_q['$gte'] = start_date
        if end_date: date_q['$lte'] = end_date + 'T23:59:59'
        q['started_at'] = date_q

    items = await db.journals.find(q, {'_id': 0}).sort('started_at', -1).to_list(limit)
    enriched = []
    for j in items:
        cls = await db.classes.find_one({'id': j.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': j.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': j.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': j.get('room_id')}, {'_id': 0, 'name': 1})
        j['class_name'] = cls.get('name') if cls else None
        j['subject_name'] = sub.get('name') if sub else None
        j['subject_code'] = sub.get('code') if sub else None
        j['teacher_name'] = teacher.get('full_name') if teacher else None
        j['room_name'] = room.get('name') if room else None
        enriched.append(serialize_doc(j))

    # Compute summary
    total_hadir = sum(j.get('siswa_hadir', 0) for j in enriched)
    total_sakit = sum(j.get('siswa_sakit', 0) for j in enriched)
    total_izin = sum(j.get('siswa_izin', 0) for j in enriched)
    total_alpa = sum(j.get('siswa_tidak_hadir', 0) for j in enriched)
    return {
        'items': enriched,
        'total': len(enriched),
        'summary': {
            'total_hadir': total_hadir, 'total_sakit': total_sakit,
            'total_izin': total_izin, 'total_alpa': total_alpa,
            'total_siswa_per_jurnal': total_hadir + total_sakit + total_izin + total_alpa,
        }
    }


@api_router.get("/admin/jurnal/stats-by-teacher")
async def admin_jurnal_stats_teacher(user: Dict = Depends(require_role('admin'))):
    """Aggregate jurnal count per guru"""
    ay = await get_active_academic_year()
    pipeline = []
    if ay:
        pipeline.append({'$match': {'academic_year_id': ay['id']}})
    pipeline.append({'$group': {'_id': '$teacher_id', 'count': {'$sum': 1}}})
    pipeline.append({'$sort': {'count': -1}})
    results = await db.journals.aggregate(pipeline).to_list(200)
    enriched = []
    for r in results:
        teacher = await db.users.find_one({'id': r['_id']}, {'_id': 0, 'full_name': 1, 'username': 1})
        enriched.append({
            'teacher_id': r['_id'],
            'teacher_name': teacher.get('full_name') if teacher else 'Unknown',
            'username': teacher.get('username') if teacher else None,
            'count': r['count'],
        })
    return enriched


# ============================================================
# EXCEL TEMPLATES & IMPORTS (Phase 4)
# ============================================================
from excel_io import (
    user_template, parse_user_rows,
    class_template, parse_class_rows,
    room_template, parse_room_rows,
    subject_template, parse_subject_rows,
    student_template, parse_student_rows,
)


def _stream_xlsx(content: bytes, filename: str):
    return StreamingResponse(
        io.BytesIO(content),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@api_router.get("/users/excel-template")
async def users_template(user: Dict = Depends(require_role('admin'))):
    return _stream_xlsx(user_template(), 'template_pengguna_matsandatama.xlsx')


@api_router.post("/users/import-excel")
async def users_import(file: UploadFile = File(...), request: Request = None,
                       user: Dict = Depends(require_role('admin'))):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya file .xlsx yang didukung")
    contents = await file.read()
    try:
        rows = parse_user_rows(contents)
    except Exception as e:
        raise HTTPException(400, f"Gagal membaca Excel: {e}")
    success = 0
    errors = []
    new_docs = []
    classes_map = {c['name']: c['id'] for c in await db.classes.find({}, {'_id': 0}).to_list(500)}
    valid_roles = set(ROLES)
    for r in rows:
        try:
            if not r['username'] or not r['password'] or not r['full_name']:
                errors.append(f"Baris {r['_row']}: username/password/nama wajib"); continue
            if not r['roles']:
                errors.append(f"Baris {r['_row']}: roles wajib"); continue
            invalid = [x for x in r['roles'] if x not in valid_roles]
            if invalid:
                errors.append(f"Baris {r['_row']}: roles tidak valid {invalid}"); continue
            existing = await db.users.find_one({'username': r['username']})
            if existing:
                errors.append(f"Baris {r['_row']}: username '{r['username']}' sudah ada"); continue
            student_class_id = None
            homeroom_class_id = None
            if 'siswa' in r['roles'] and r.get('kelas_siswa'):
                student_class_id = classes_map.get(r['kelas_siswa'])
                if not student_class_id:
                    errors.append(f"Baris {r['_row']}: kelas '{r['kelas_siswa']}' tidak ditemukan"); continue
            if 'wali_kelas' in r['roles'] and r.get('wali_kelas'):
                homeroom_class_id = classes_map.get(r['wali_kelas'])
                if not homeroom_class_id:
                    errors.append(f"Baris {r['_row']}: wali_kelas '{r['wali_kelas']}' tidak ditemukan"); continue
            u = UserModel(
                username=r['username'],
                password_hash=hash_password(r['password']),
                full_name=r['full_name'],
                roles=r['roles'],
                nip_nuptk=r.get('nip_nuptk'), nisn=r.get('nisn'),
                email=r.get('email'), phone=r.get('phone'), gender=r.get('gender'),
                student_class_id=student_class_id,
                homeroom_class_id=homeroom_class_id,
            )
            doc = u.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            new_docs.append(doc)
            success += 1
            # Update wali_kelas reference in class
            if homeroom_class_id:
                await db.classes.update_one({'id': homeroom_class_id}, {'$set': {'homeroom_teacher_id': u.id}})
        except Exception as e:
            errors.append(f"Baris {r['_row']}: {e}")
    if new_docs:
        await db.users.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'user', None,
                   details={'success': success, 'errors': len(errors), 'filename': file.filename}, request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


@api_router.get("/classes/excel-template")
async def classes_template_dl(user: Dict = Depends(require_role('admin'))):
    return _stream_xlsx(class_template(), 'template_kelas_matsandatama.xlsx')


@api_router.post("/classes/import-excel")
async def classes_import(file: UploadFile = File(...), request: Request = None,
                          user: Dict = Depends(require_role('admin'))):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya .xlsx")
    contents = await file.read()
    rows = parse_class_rows(contents)
    ay = await get_active_academic_year()
    if not ay:
        raise HTTPException(400, "Tidak ada TP aktif")
    users_map = {u['username']: u['id'] for u in await db.users.find({}, {'_id': 0, 'username': 1, 'id': 1}).to_list(2000)}
    rooms_map = {r['name']: r['id'] for r in await db.rooms.find({}, {'_id': 0, 'name': 1, 'id': 1}).to_list(500)}
    success = 0
    errors = []
    new_docs = []
    for r in rows:
        try:
            if not r['name']:
                errors.append(f"Baris {r['_row']}: nama kelas wajib"); continue
            existing = await db.classes.find_one({'name': r['name'], 'academic_year_id': ay['id']})
            if existing:
                errors.append(f"Baris {r['_row']}: kelas '{r['name']}' sudah ada di TP aktif"); continue
            homeroom_id = users_map.get(r.get('wali_kelas_username')) if r.get('wali_kelas_username') else None
            room_id = rooms_map.get(r.get('ruang_kode')) if r.get('ruang_kode') else None
            cls = ClassModel(
                name=r['name'], grade=r['grade'], parallel=r['parallel'],
                academic_year_id=ay['id'],
                homeroom_teacher_id=homeroom_id, room_id=room_id,
                is_accelerated=r.get('is_accelerated', False),
            )
            doc = cls.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            new_docs.append(doc)
            success += 1
        except Exception as e:
            errors.append(f"Baris {r['_row']}: {e}")
    if new_docs:
        await db.classes.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'class', None, details={'success': success, 'errors': len(errors)}, request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


@api_router.get("/rooms/excel-template")
async def rooms_template_dl(user: Dict = Depends(require_role('admin'))):
    return _stream_xlsx(room_template(), 'template_ruangan_matsandatama.xlsx')


@api_router.post("/rooms/import-excel")
async def rooms_import(file: UploadFile = File(...), request: Request = None,
                       user: Dict = Depends(require_role('admin'))):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya .xlsx")
    contents = await file.read()
    rows = parse_room_rows(contents)
    success = 0
    errors = []
    new_docs = []
    for r in rows:
        try:
            if not r['name']:
                errors.append(f"Baris {r['_row']}: kode ruang wajib"); continue
            existing = await db.rooms.find_one({'name': r['name']})
            if existing:
                errors.append(f"Baris {r['_row']}: kode '{r['name']}' sudah ada"); continue
            if r.get('qr_mode') == 'dynamic':
                r['qr_secret'] = pyotp.random_base32()
            rm = RoomModel(**{k: v for k, v in r.items() if not k.startswith('_')})
            doc = rm.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            new_docs.append(doc)
            success += 1
        except Exception as e:
            errors.append(f"Baris {r['_row']}: {e}")
    if new_docs:
        await db.rooms.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'room', None, details={'success': success, 'errors': len(errors)}, request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


@api_router.get("/subjects/excel-template")
async def subjects_template_dl(user: Dict = Depends(require_role('admin'))):
    return _stream_xlsx(subject_template(), 'template_mapel_matsandatama.xlsx')


@api_router.post("/subjects/import-excel")
async def subjects_import(file: UploadFile = File(...), request: Request = None,
                          user: Dict = Depends(require_role('admin'))):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya .xlsx")
    contents = await file.read()
    rows = parse_subject_rows(contents)
    success = 0
    errors = []
    new_docs = []
    for r in rows:
        try:
            if not r['code'] or not r['name']:
                errors.append(f"Baris {r['_row']}: kode & nama wajib"); continue
            existing = await db.subjects.find_one({'code': r['code']})
            if existing:
                errors.append(f"Baris {r['_row']}: kode '{r['code']}' sudah ada"); continue
            s = SubjectModel(code=r['code'], name=r['name'])
            doc = s.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            new_docs.append(doc)
            success += 1
        except Exception as e:
            errors.append(f"Baris {r['_row']}: {e}")
    if new_docs:
        await db.subjects.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'subject', None, details={'success': success, 'errors': len(errors)}, request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


@api_router.get("/students/excel-template")
async def students_template_dl(user: Dict = Depends(require_role('admin'))):
    return _stream_xlsx(student_template(), 'template_siswa_matsandatama.xlsx')


@api_router.post("/students/import-excel")
async def students_import(file: UploadFile = File(...), request: Request = None,
                          user: Dict = Depends(require_role('admin'))):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya .xlsx")
    contents = await file.read()
    rows = parse_student_rows(contents)
    classes_map = {c['name']: c['id'] for c in await db.classes.find({}, {'_id': 0}).to_list(500)}
    success = 0
    errors = []
    new_docs = []
    for r in rows:
        try:
            if not all([r['username'], r['password'], r['full_name'], r['nisn'], r['kelas']]):
                errors.append(f"Baris {r['_row']}: username/password/nama/NISN/kelas wajib"); continue
            cls_id = classes_map.get(r['kelas'])
            if not cls_id:
                errors.append(f"Baris {r['_row']}: kelas '{r['kelas']}' tidak ditemukan"); continue
            existing = await db.users.find_one({'username': r['username']})
            if existing:
                errors.append(f"Baris {r['_row']}: username '{r['username']}' sudah ada"); continue
            u = UserModel(
                username=r['username'], password_hash=hash_password(r['password']),
                full_name=r['full_name'], roles=['siswa'], nisn=r['nisn'],
                gender=r.get('gender'), student_class_id=cls_id,
                birth_place=r.get('birth_place'), birth_date=r.get('birth_date'),
                address=r.get('address'), email=r.get('email'), phone=r.get('phone'),
            )
            doc = u.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            new_docs.append(doc)
            success += 1
        except Exception as e:
            errors.append(f"Baris {r['_row']}: {e}")
    if new_docs:
        await db.users.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'student', None, details={'success': success, 'errors': len(errors)}, request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


# ============================================================
# SMTP & PASSWORD RESET (Phase 4)
# ============================================================
from email_utils import (
    create_reset_token, validate_reset_token, consume_reset_token,
    send_email, build_reset_email,
)


@api_router.post("/admin/settings/test-smtp")
async def test_smtp(payload: Dict[str, Any], request: Request = None,
                    user: Dict = Depends(require_role('admin'))):
    """Send test email to verify SMTP configuration"""
    settings = await get_settings()
    # Allow overriding for test
    cfg = {**settings, **payload}
    to_email = payload.get('to_email') or user.get('email')
    if not to_email:
        raise HTTPException(400, "Email tujuan wajib (set via payload to_email)")
    result = send_email(
        cfg, to_email,
        subject="Test SMTP - Super Apps MATSANDATAMA",
        body_text="Selamat! Konfigurasi SMTP Anda berhasil. Email ini dikirim sebagai uji coba.",
        body_html="<p>Selamat! Konfigurasi SMTP Anda <strong>berhasil</strong>. Email ini dikirim sebagai uji coba dari Super Apps MATSANDATAMA.</p>",
    )
    await log_audit(user, 'test_smtp', 'settings', None, details={'success': result['success']}, request=request)
    return result


class ForgotPasswordRequest(BaseModel):
    identifier: str  # username OR email


class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str


@api_router.post("/auth/forgot-password")
async def forgot_password(req: ForgotPasswordRequest, request: Request):
    """Send password reset email. Always returns OK to prevent enumeration."""
    user = await db.users.find_one({'$or': [
        {'username': req.identifier},
        {'email': req.identifier},
    ]})
    if not user or not user.get('email') or not user.get('is_active', True):
        # Silent fail to prevent username enumeration
        return {'message': 'Jika akun terdaftar dengan email, instruksi reset telah dikirim.'}
    settings = await get_settings()
    if not settings.get('smtp_host'):
        return {'message': 'Fitur reset email belum tersedia. Hubungi admin.'}
    token = create_reset_token(user['id'], user['email'])
    base = settings.get('app_public_url') or str(request.base_url).rstrip('/')
    reset_link = f"{base}/reset-password?token={token}"
    body = build_reset_email(reset_link, user['username'],
                             settings.get('app_name', 'Super Apps MATSANDATAMA'),
                             settings.get('school_name', 'MTsN 2 Kota Malang'))
    send_result = send_email(
        settings, user['email'],
        subject=f"Reset Password - {settings.get('app_name', 'Super Apps MATSANDATAMA')}",
        body_text=body['text'], body_html=body['html'],
    )
    await log_security('forgot_password',
                       user.get('username'),
                       {'sent': send_result.get('success'), 'email': user['email']},
                       request)
    return {'message': 'Jika akun terdaftar dengan email, instruksi reset telah dikirim.'}


@api_router.get("/auth/reset-password/validate/{token}")
async def reset_password_validate(token: str):
    item = validate_reset_token(token)
    if not item:
        raise HTTPException(400, "Token tidak valid atau kedaluwarsa")
    user = await db.users.find_one({'id': item['user_id']}, {'_id': 0, 'username': 1, 'email': 1})
    return {'valid': True, 'username': user.get('username') if user else None}


@api_router.post("/auth/reset-password")
async def reset_password(req: ResetPasswordRequest, request: Request):
    item = consume_reset_token(req.token)
    if not item:
        raise HTTPException(400, "Token tidak valid atau kedaluwarsa")
    if len(req.new_password) < 6:
        raise HTTPException(400, "Password minimal 6 karakter")
    user = await db.users.find_one({'id': item['user_id']})
    if not user:
        raise HTTPException(404, "User tidak ditemukan")
    await db.users.update_one({'id': user['id']},
                               {'$set': {'password_hash': hash_password(req.new_password)}})
    await log_security('password_reset', user.get('username'), {'method': 'email_token'}, request)
    return {'message': 'Password berhasil direset. Silakan login.'}


# ============================================================
# PRESTASI SISWA (Achievement)
# ============================================================
from models_phase4 import (
    StudentAchievementModel, ExtracurricularModel, ExtracurricularMemberModel,
    ExtracurricularAttendanceModel, ExtracurricularGradeModel, GradeEntryModel,
)


@api_router.get("/achievements")
async def list_achievements(student_id: Optional[str] = None,
                             only_verified: bool = False,
                             user: Dict = Depends(get_current_user)):
    q = {}
    if student_id:
        q['student_id'] = student_id
    if only_verified:
        q['is_verified'] = True
    # Permission: students can only see their own; admin/wk see all
    is_admin = 'admin' in user.get('roles', [])
    if not is_admin and 'wali_kelas' not in user.get('roles', []):
        if student_id and student_id != user['id'] and user['id'] not in [user['id']]:
            # Allow seeing own
            if user['id'] != student_id:
                raise HTTPException(403, "Tidak diizinkan")
        if not student_id:
            q['student_id'] = user['id']
    items = await db.achievements.find(q, {'_id': 0}).sort('date', -1).to_list(500)
    enriched = []
    for a in items:
        s = await db.users.find_one({'id': a.get('student_id')}, {'_id': 0, 'full_name': 1, 'nisn': 1, 'student_class_id': 1})
        if s:
            a['student_name'] = s.get('full_name')
            a['student_nisn'] = s.get('nisn')
            cls = await db.classes.find_one({'id': s.get('student_class_id')}, {'_id': 0, 'name': 1})
            a['class_name'] = cls.get('name') if cls else None
        if a.get('verified_by'):
            v = await db.users.find_one({'id': a['verified_by']}, {'_id': 0, 'full_name': 1})
            a['verifier_name'] = v.get('full_name') if v else None
        enriched.append(serialize_doc(a))
    return enriched


@api_router.post("/achievements")
async def create_achievement(payload: Dict, request: Request, user: Dict = Depends(get_current_user)):
    # Siswa can self-submit; admin can submit for anyone
    target_student = payload.get('student_id') or user['id']
    if 'siswa' in user.get('roles', []) and target_student != user['id']:
        raise HTTPException(403, "Siswa hanya bisa input prestasi sendiri")
    a = StudentAchievementModel(
        student_id=target_student,
        name=payload.get('name', ''),
        category=payload.get('category'),
        level=payload.get('level'),
        rank=payload.get('rank'),
        organizer=payload.get('organizer'),
        date=payload.get('date'),
        description=payload.get('description'),
        certificate_url=payload.get('certificate_url'),
        submitted_by=user['id'],
    )
    doc = a.model_dump()
    doc['submitted_at'] = doc['submitted_at'].isoformat()
    await db.achievements.insert_one(doc)
    await log_audit(user, 'create', 'achievement', a.id, details={'name': a.name}, request=request)
    return serialize_doc(doc)


@api_router.put("/achievements/{aid}")
async def update_achievement(aid: str, payload: Dict, request: Request,
                              user: Dict = Depends(get_current_user)):
    existing = await db.achievements.find_one({'id': aid})
    if not existing:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_owner = existing.get('submitted_by') == user['id'] or existing.get('student_id') == user['id']
    if not (is_admin or is_owner):
        raise HTTPException(403, "Tidak diizinkan")
    payload.pop('_id', None); payload.pop('id', None)
    await db.achievements.update_one({'id': aid}, {'$set': payload})
    await log_audit(user, 'update', 'achievement', aid, request=request)
    doc = await db.achievements.find_one({'id': aid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.put("/achievements/{aid}/verify")
async def verify_achievement(aid: str, request: Request,
                              user: Dict = Depends(require_role('admin', 'wali_kelas'))):
    await db.achievements.update_one({'id': aid}, {'$set': {
        'is_verified': True, 'verified_by': user['id'],
        'verified_at': datetime.utcnow().isoformat(),
    }})
    await log_audit(user, 'verify', 'achievement', aid, request=request)
    doc = await db.achievements.find_one({'id': aid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/achievements/{aid}")
async def delete_achievement(aid: str, request: Request, user: Dict = Depends(get_current_user)):
    existing = await db.achievements.find_one({'id': aid})
    if not existing:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_owner = existing.get('submitted_by') == user['id']
    if not (is_admin or is_owner):
        raise HTTPException(403, "Tidak diizinkan")
    await db.achievements.delete_one({'id': aid})
    await log_audit(user, 'delete', 'achievement', aid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# EKSTRAKURIKULER (Phase 4)
# ============================================================
@api_router.get("/extracurriculars")
async def list_extras(user: Dict = Depends(get_current_user)):
    items = await db.extracurriculars.find({}, {'_id': 0}).sort('name', 1).to_list(200)
    enriched = []
    for e in items:
        coach = await db.users.find_one({'id': e.get('coach_id')}, {'_id': 0, 'full_name': 1})
        e['coach_name'] = coach.get('full_name') if coach else None
        member_count = await db.extracurricular_members.count_documents({'extracurricular_id': e['id'], 'is_active': True})
        e['member_count'] = member_count
        enriched.append(serialize_doc(e))
    return enriched


@api_router.post("/extracurriculars")
async def create_extra(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    ay = await get_active_academic_year()
    payload['academic_year_id'] = ay['id'] if ay else None
    e = ExtracurricularModel(**payload)
    doc = e.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.extracurriculars.insert_one(doc)
    await log_audit(user, 'create', 'extracurricular', e.id, details={'name': e.name}, request=request)
    return serialize_doc(doc)


@api_router.put("/extracurriculars/{eid}")
async def update_extra(eid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    payload.pop('_id', None); payload.pop('id', None)
    await db.extracurriculars.update_one({'id': eid}, {'$set': payload})
    await log_audit(user, 'update', 'extracurricular', eid, request=request)
    doc = await db.extracurriculars.find_one({'id': eid}, {'_id': 0})
    return serialize_doc(doc)


@api_router.delete("/extracurriculars/{eid}")
async def delete_extra(eid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.extracurriculars.delete_one({'id': eid})
    await db.extracurricular_members.delete_many({'extracurricular_id': eid})
    await log_audit(user, 'delete', 'extracurricular', eid, request=request)
    return {'message': 'Dihapus'}


@api_router.get("/extracurriculars/{eid}/members")
async def list_extra_members(eid: str, user: Dict = Depends(get_current_user)):
    members = await db.extracurricular_members.find({'extracurricular_id': eid}, {'_id': 0}).to_list(500)
    enriched = []
    for m in members:
        s = await db.users.find_one({'id': m.get('student_id')}, {'_id': 0, 'full_name': 1, 'nisn': 1, 'student_class_id': 1})
        if s:
            cls = await db.classes.find_one({'id': s.get('student_class_id')}, {'_id': 0, 'name': 1})
            m['student_name'] = s.get('full_name')
            m['student_nisn'] = s.get('nisn')
            m['class_name'] = cls.get('name') if cls else None
        enriched.append(serialize_doc(m))
    return enriched


@api_router.post("/extracurriculars/{eid}/members")
async def add_extra_members(eid: str, payload: Dict, request: Request,
                             user: Dict = Depends(get_current_user)):
    # admin and guru_ekstrakurikuler (coach) can manage
    extra = await db.extracurriculars.find_one({'id': eid})
    if not extra:
        raise HTTPException(404, "Ekskul tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_coach = extra.get('coach_id') == user['id']
    if not (is_admin or is_coach):
        raise HTTPException(403, "Tidak diizinkan")
    student_ids = payload.get('student_ids', [])
    inserted = 0
    for sid in student_ids:
        existing = await db.extracurricular_members.find_one({'extracurricular_id': eid, 'student_id': sid})
        if existing:
            await db.extracurricular_members.update_one({'_id': existing['_id']}, {'$set': {'is_active': True}})
            continue
        m = ExtracurricularMemberModel(extracurricular_id=eid, student_id=sid)
        doc = m.model_dump()
        doc['joined_at'] = doc['joined_at'].isoformat()
        await db.extracurricular_members.insert_one(doc)
        inserted += 1
    await log_audit(user, 'add_members', 'extracurricular', eid, details={'count': len(student_ids)}, request=request)
    return {'inserted': inserted, 'total': len(student_ids)}


@api_router.delete("/extracurriculars/{eid}/members/{mid}")
async def remove_extra_member(eid: str, mid: str, request: Request,
                                user: Dict = Depends(get_current_user)):
    extra = await db.extracurriculars.find_one({'id': eid})
    is_admin = 'admin' in user.get('roles', [])
    is_coach = extra and extra.get('coach_id') == user['id']
    if not (is_admin or is_coach):
        raise HTTPException(403, "Tidak diizinkan")
    await db.extracurricular_members.delete_one({'id': mid})
    return {'message': 'Dihapus'}


@api_router.post("/extracurriculars/{eid}/attendance")
async def submit_extra_attendance(eid: str, payload: Dict, request: Request,
                                    user: Dict = Depends(get_current_user)):
    extra = await db.extracurriculars.find_one({'id': eid})
    is_admin = 'admin' in user.get('roles', [])
    is_coach = extra and extra.get('coach_id') == user['id']
    if not (is_admin or is_coach):
        raise HTTPException(403, "Tidak diizinkan")
    date = payload.get('date')
    records = payload.get('records', [])
    summary = {'hadir': 0, 'sakit': 0, 'izin': 0, 'alpa': 0}
    for r in records:
        st = r.get('status', 'hadir')
        summary[st] = summary.get(st, 0) + 1
    existing = await db.extracurricular_attendance.find_one({'extracurricular_id': eid, 'date': date})
    doc = {
        'extracurricular_id': eid, 'date': date, 'records': records,
        'recorded_by': user['id'], 'recorded_at': datetime.utcnow().isoformat(),
        'summary': summary,
    }
    if existing:
        await db.extracurricular_attendance.update_one({'_id': existing['_id']}, {'$set': doc})
        doc['id'] = existing.get('id', str(uuid.uuid4()))
    else:
        doc['id'] = str(uuid.uuid4())
        await db.extracurricular_attendance.insert_one(doc)
    await log_audit(user, 'attendance', 'extracurricular', eid, details={'date': date, 'summary': summary}, request=request)
    return serialize_doc(doc)


@api_router.get("/extracurriculars/{eid}/attendance")
async def get_extra_attendance(eid: str, user: Dict = Depends(get_current_user)):
    items = await db.extracurricular_attendance.find({'extracurricular_id': eid}, {'_id': 0}).sort('date', -1).to_list(200)
    return [serialize_doc(i) for i in items]


@api_router.post("/extracurriculars/{eid}/grades")
async def submit_extra_grades(eid: str, payload: Dict, request: Request,
                                user: Dict = Depends(get_current_user)):
    extra = await db.extracurriculars.find_one({'id': eid})
    is_admin = 'admin' in user.get('roles', [])
    is_coach = extra and extra.get('coach_id') == user['id']
    if not (is_admin or is_coach):
        raise HTTPException(403, "Tidak diizinkan")
    ay = await get_active_academic_year()
    semester = payload.get('semester', 'ganjil')
    grades = payload.get('grades', [])  # [{student_id, predicate, description}]
    inserted = 0
    for g in grades:
        existing = await db.extracurricular_grades.find_one({
            'extracurricular_id': eid, 'student_id': g['student_id'],
            'academic_year_id': ay['id'] if ay else None, 'semester': semester,
        })
        doc = {
            'extracurricular_id': eid, 'student_id': g['student_id'],
            'academic_year_id': ay['id'] if ay else None, 'semester': semester,
            'predicate': g.get('predicate'), 'description': g.get('description'),
            'submitted_by': user['id'], 'submitted_at': datetime.utcnow().isoformat(),
        }
        if existing:
            await db.extracurricular_grades.update_one({'_id': existing['_id']}, {'$set': doc})
        else:
            doc['id'] = str(uuid.uuid4())
            await db.extracurricular_grades.insert_one(doc)
        inserted += 1
    await log_audit(user, 'grades', 'extracurricular', eid, details={'count': inserted, 'semester': semester}, request=request)
    return {'success': inserted}


@api_router.get("/extracurriculars/{eid}/grades")
async def get_extra_grades(eid: str, semester: Optional[str] = None, user: Dict = Depends(get_current_user)):
    q = {'extracurricular_id': eid}
    if semester: q['semester'] = semester
    items = await db.extracurricular_grades.find(q, {'_id': 0}).to_list(500)
    return [serialize_doc(i) for i in items]


# ============================================================
# E-RAPOR (Grade entries)
# ============================================================
@api_router.post("/grades/bulk")
async def submit_grades_bulk(payload: Dict, request: Request, user: Dict = Depends(get_current_user)):
    """Bulk submit grades for one class+subject+semester."""
    ay = await get_active_academic_year()
    class_id = payload.get('class_id')
    subject_id = payload.get('subject_id')
    semester = payload.get('semester', 'ganjil')
    entries = payload.get('entries', [])  # [{student_id, nilai_pengetahuan, nilai_keterampilan, predicate, description}]

    # Permission: admin or teacher of that subject/class
    is_admin = 'admin' in user.get('roles', [])
    if not is_admin:
        sched = await db.schedules.find_one({
            'class_id': class_id, 'subject_id': subject_id,
            'teacher_id': user['id'], 'academic_year_id': ay['id'] if ay else None,
        })
        if not sched:
            raise HTTPException(403, "Anda bukan pengampu mapel ini di kelas tersebut")

    inserted = 0
    for e in entries:
        nilai_p = e.get('nilai_pengetahuan')
        nilai_k = e.get('nilai_keterampilan')
        nilai_akhir = None
        if nilai_p is not None and nilai_k is not None:
            try:
                nilai_akhir = (float(nilai_p) + float(nilai_k)) / 2
            except Exception:
                nilai_akhir = None
        elif nilai_p is not None:
            nilai_akhir = float(nilai_p)
        predicate = e.get('predicate')
        if not predicate and nilai_akhir is not None:
            if nilai_akhir >= 88: predicate = 'A'
            elif nilai_akhir >= 76: predicate = 'B'
            elif nilai_akhir >= 60: predicate = 'C'
            else: predicate = 'D'
        existing = await db.grade_entries.find_one({
            'student_id': e['student_id'], 'class_id': class_id,
            'subject_id': subject_id, 'semester': semester,
            'academic_year_id': ay['id'] if ay else None,
        })
        doc = {
            'student_id': e['student_id'], 'class_id': class_id,
            'subject_id': subject_id, 'teacher_id': user['id'],
            'academic_year_id': ay['id'] if ay else None, 'semester': semester,
            'nilai_pengetahuan': float(nilai_p) if nilai_p is not None else None,
            'nilai_keterampilan': float(nilai_k) if nilai_k is not None else None,
            'nilai_akhir': nilai_akhir, 'predicate': predicate,
            'description': e.get('description'),
            'submitted_by': user['id'], 'submitted_at': datetime.utcnow().isoformat(),
        }
        if existing:
            await db.grade_entries.update_one({'_id': existing['_id']}, {'$set': doc})
        else:
            doc['id'] = str(uuid.uuid4())
            await db.grade_entries.insert_one(doc)
        inserted += 1
    await log_audit(user, 'grades_bulk', 'grade_entries', None,
                   details={'class_id': class_id, 'subject_id': subject_id, 'count': inserted}, request=request)
    return {'success': inserted}


@api_router.get("/grades")
async def list_grades(class_id: Optional[str] = None, subject_id: Optional[str] = None,
                       student_id: Optional[str] = None, semester: Optional[str] = None,
                       academic_year_id: Optional[str] = None,
                       user: Dict = Depends(get_current_user)):
    q = {}
    if class_id: q['class_id'] = class_id
    if subject_id: q['subject_id'] = subject_id
    if student_id: q['student_id'] = student_id
    if semester: q['semester'] = semester
    if academic_year_id: q['academic_year_id'] = academic_year_id
    # Students can only see own
    if 'admin' not in user.get('roles', []) and 'siswa' in user.get('roles', []):
        q['student_id'] = user['id']
    items = await db.grade_entries.find(q, {'_id': 0}).to_list(2000)
    enriched = []
    for g in items:
        s = await db.users.find_one({'id': g.get('student_id')}, {'_id': 0, 'full_name': 1, 'nisn': 1})
        sub = await db.subjects.find_one({'id': g.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        if s: g['student_name'] = s.get('full_name'); g['student_nisn'] = s.get('nisn')
        if sub: g['subject_name'] = sub.get('name'); g['subject_code'] = sub.get('code')
        enriched.append(serialize_doc(g))
    return enriched


@api_router.get("/grades/rapor/{student_id}")
async def get_student_rapor(student_id: str, semester: Optional[str] = None,
                              user: Dict = Depends(get_current_user)):
    """Get rapor view for a student"""
    student = await db.users.find_one({'id': student_id}, {'_id': 0, 'password_hash': 0})
    if not student:
        raise HTTPException(404, "Siswa tidak ditemukan")
    # Permission check
    is_admin = 'admin' in user.get('roles', [])
    is_self = user['id'] == student_id
    cls = await db.classes.find_one({'id': student.get('student_class_id')}, {'_id': 0})
    is_homeroom = cls and cls.get('homeroom_teacher_id') == user['id']
    if not (is_admin or is_self or is_homeroom):
        raise HTTPException(403, "Tidak diizinkan")
    ay = await get_active_academic_year()
    q = {'student_id': student_id}
    if semester: q['semester'] = semester
    if ay: q['academic_year_id'] = ay['id']
    items = await db.grade_entries.find(q, {'_id': 0}).to_list(200)
    enriched = []
    for g in items:
        sub = await db.subjects.find_one({'id': g.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': g.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        if sub: g['subject_name'] = sub.get('name'); g['subject_code'] = sub.get('code')
        if teacher: g['teacher_name'] = teacher.get('full_name')
        enriched.append(serialize_doc(g))
    return {
        'student': serialize_doc(student),
        'class': serialize_doc(cls) if cls else None,
        'academic_year': ay,
        'grades': enriched,
        'average': round(sum(g.get('nilai_akhir', 0) or 0 for g in enriched) / len(enriched), 2) if enriched else 0,
    }


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"], allow_headers=["*"],
)


@app.on_event("startup")
async def startup_event():
    try:
        from seed_data import seed_all, refresh_demo_schedule
        await seed_all(db)
        await refresh_demo_schedule(db)
    except Exception as e:
        logger.error(f"Seed error: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
