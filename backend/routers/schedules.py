"""Schedules: CRUD + workflow (draft/submit/lock/unlock) + grid + my-today + Excel import + Piket schedules CRUD + Conflict detection."""
import io
import uuid
from datetime import datetime
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse

from core import (
    db,
    get_active_academic_year,
    get_current_user,
    get_settings,
    log_audit,
    logger,
    require_role,
    serialize_doc,
)
from journal_core import current_day_id, now_wib
from models import ScheduleModel

router = APIRouter()


# ============================================================
# CONFLICT DETECTION (Phase 6)
# ============================================================
def _times_overlap(start_a: str, end_a: str, start_b: str, end_b: str) -> bool:
    """Two HH:MM ranges overlap? Inclusive boundary touches are NOT considered overlap.
    Example: 07:00-07:45 vs 07:45-08:30 = no overlap.
    """
    try:
        return start_a < end_b and start_b < end_a
    except Exception:
        return False


async def _find_conflicts(
    academic_year_id: str,
    day: str,
    start_time: str,
    end_time: str,
    teacher_id: Optional[str] = None,
    room_id: Optional[str] = None,
    class_id: Optional[str] = None,
    exclude_id: Optional[str] = None,
) -> Dict[str, List[Dict]]:
    """Find conflicts grouped by type. Returns:
    { 'teacher': [...], 'room': [...], 'class': [...] }
    Each entry includes: id, day, start_time, end_time, teacher_name, subject_name,
    class_name, room_name, status (draft/submitted/locked).
    """
    q = {'academic_year_id': academic_year_id, 'day': day}
    if exclude_id:
        q['id'] = {'$ne': exclude_id}
    candidates = await db.schedules.find(q, {'_id': 0}).to_list(500)
    overlapping = [s for s in candidates if _times_overlap(start_time, end_time,
                                                            s.get('start_time', ''),
                                                            s.get('end_time', ''))]

    async def _enrich(s):
        out = dict(s)
        if s.get('teacher_id'):
            t = await db.users.find_one({'id': s['teacher_id']}, {'_id': 0, 'full_name': 1})
            out['teacher_name'] = t.get('full_name') if t else None
        if s.get('subject_id'):
            sub = await db.subjects.find_one({'id': s['subject_id']}, {'_id': 0, 'name': 1, 'code': 1})
            if sub:
                out['subject_name'] = sub.get('name')
                out['subject_code'] = sub.get('code')
        if s.get('class_id'):
            cls = await db.classes.find_one({'id': s['class_id']}, {'_id': 0, 'name': 1})
            out['class_name'] = cls.get('name') if cls else None
        if s.get('room_id'):
            r = await db.rooms.find_one({'id': s['room_id']}, {'_id': 0, 'name': 1})
            out['room_name'] = r.get('name') if r else None
        return serialize_doc(out)

    result = {'teacher': [], 'room': [], 'class': []}
    for s in overlapping:
        enriched = await _enrich(s)
        if teacher_id and s.get('teacher_id') == teacher_id:
            result['teacher'].append(enriched)
        if room_id and s.get('room_id') == room_id:
            result['room'].append(enriched)
        if class_id and s.get('class_id') == class_id:
            result['class'].append(enriched)
    return result


def _conflict_message(conflicts: Dict[str, List[Dict]]) -> str:
    parts = []
    for c in conflicts.get('teacher', []):
        parts.append(f"Guru '{c.get('teacher_name', '-')}' sudah ada di {c.get('class_name', '-')} "
                     f"({c.get('subject_name', '-')}) {c.get('start_time')}-{c.get('end_time')} [{c.get('status', '-')}]")
    for c in conflicts.get('room', []):
        parts.append(f"Ruang '{c.get('room_name', '-')}' sudah dipakai {c.get('teacher_name', '-')} "
                     f"({c.get('class_name', '-')}) {c.get('start_time')}-{c.get('end_time')} [{c.get('status', '-')}]")
    return ' | '.join(parts) if parts else ''


@router.get("/schedules/conflict-check")
async def schedules_conflict_check(
    day: str, start_time: str, end_time: str,
    academic_year_id: Optional[str] = None,
    teacher_id: Optional[str] = None,
    room_id: Optional[str] = None,
    class_id: Optional[str] = None,
    exclude_id: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """Cek konflik jadwal: bentrok guru / ruang / kelas pada hari & jam tertentu.

    Returns: { has_conflict: bool, teacher: [...], room: [...], class: [...], message: str }
    """
    ay_id = academic_year_id
    if not ay_id:
        ay = await get_active_academic_year()
        if not ay:
            return {'has_conflict': False, 'teacher': [], 'room': [], 'class': [], 'message': ''}
        ay_id = ay['id']
    conflicts = await _find_conflicts(ay_id, day, start_time, end_time,
                                       teacher_id=teacher_id, room_id=room_id, class_id=class_id,
                                       exclude_id=exclude_id)
    has = bool(conflicts['teacher'] or conflicts['room'])  # only guru + ruang BLOCK
    return {
        'has_conflict': has,
        'teacher': conflicts['teacher'],
        'room': conflicts['room'],
        'class': conflicts['class'],  # informational only
        'message': _conflict_message(conflicts),
    }


# ============================================================
# SCHEDULES LIST/CREATE
# ============================================================

# More specific routes must come BEFORE generic routes in FastAPI
@router.get("/schedules/grouped")
async def list_schedules_grouped(
    academic_year_id: Optional[str] = None, semester: Optional[str] = None,
    class_id: Optional[str] = None, teacher_id: Optional[str] = None, day: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """
    Get schedules grouped by JTM (Jam Tugas Mengajar).
    Consecutive teaching hours in same class, day, subject are grouped together.
    """
    q = {}
    if academic_year_id: q['academic_year_id'] = academic_year_id
    if semester: q['semester'] = semester
    if class_id: q['class_id'] = class_id
    if teacher_id: q['teacher_id'] = teacher_id
    if day: q['day'] = day

    items = await db.schedules.find(q, {'_id': 0}).sort([('day', 1), ('start_time', 1)]).to_list(2000)

    # Enrich first
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['subject_code'] = sub.get('code') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        if s.get('locked_by'):
            lb = await db.users.find_one({'id': s['locked_by']}, {'_id': 0, 'full_name': 1})
            s['locked_by_name'] = lb.get('full_name') if lb else None
        enriched.append(serialize_doc(s))

    # Group by JTM
    grouped = _group_schedules_by_jtm(enriched)

    return grouped


@router.get("/schedules")
async def list_schedules(
    academic_year_id: Optional[str] = None, semester: Optional[str] = None,
    class_id: Optional[str] = None, teacher_id: Optional[str] = None, day: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    q = {}
    if academic_year_id: q['academic_year_id'] = academic_year_id
    if semester: q['semester'] = semester
    if class_id: q['class_id'] = class_id
    if teacher_id: q['teacher_id'] = teacher_id
    if day: q['day'] = day
    items = await db.schedules.find(q, {'_id': 0}).sort([('day', 1), ('start_time', 1)]).to_list(2000)
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['subject_code'] = sub.get('code') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        if s.get('locked_by'):
            lb = await db.users.find_one({'id': s['locked_by']}, {'_id': 0, 'full_name': 1})
            s['locked_by_name'] = lb.get('full_name') if lb else None
        enriched.append(serialize_doc(s))
    return enriched


def _group_schedules_by_jtm(schedules: List[Dict]) -> List[Dict]:
    """
    Group teaching hours into JTM blocks.
    Same day, class, subject, teacher = 1 JTM entry, even if separated by break times.

    Example: Math class at hours 2, 3 (before break), then hours 4, 5 (after break)
    If same teacher, subject, class, day = counted as 4 JTM total.

    Returns list of grouped schedules with:
    - jtm_count: number of teaching hours
    - hours: list of hour numbers (e.g., [2, 3, 4, 5])
    - hour_range: string like "Jam ke-2, 3, 4, 5"
    - schedule_ids: list of original schedule IDs in this group
    """
    from datetime import datetime as dt

    # Sort by day, class, subject, teacher, then start_time
    sorted_schedules = sorted(schedules, key=lambda s: (
        s.get('day', ''),
        s.get('class_id', ''),
        s.get('subject_id', ''),
        s.get('teacher_id', ''),
        s.get('start_time', '')
    ))

    grouped = []
    i = 0

    while i < len(sorted_schedules):
        current = sorted_schedules[i]
        group = [current]
        group_ids = [current.get('id')]

        # Try to find all schedules with same day, class, subject, teacher
        # even if they're not strictly consecutive (breaks allowed)
        j = i + 1
        last_end_time = current.get('end_time')

        while j < len(sorted_schedules):
            next_schedule = sorted_schedules[j]

            # Check if same day, class, subject, teacher
            if (next_schedule.get('day') == current.get('day') and
                next_schedule.get('class_id') == current.get('class_id') and
                next_schedule.get('subject_id') == current.get('subject_id') and
                next_schedule.get('teacher_id') == current.get('teacher_id')):

                # Allow small gaps for break times (up to 2 hours gap)
                # Parse time strings to check gap
                try:
                    from datetime import datetime
                    last_end = datetime.strptime(last_end_time, '%H:%M')
                    next_start = datetime.strptime(next_schedule.get('start_time'), '%H:%M')
                    gap_minutes = (next_start - last_end).total_seconds() / 60

                    # If gap is reasonable (< 120 minutes), it's likely just a break
                    # and we should group them together
                    if gap_minutes <= 120:
                        group.append(next_schedule)
                        group_ids.append(next_schedule.get('id'))
                        last_end_time = next_schedule.get('end_time')
                        j += 1
                    else:
                        break
                except:
                    # If time parsing fails, use strict consecutive check
                    if last_end_time == next_schedule.get('start_time'):
                        group.append(next_schedule)
                        group_ids.append(next_schedule.get('id'))
                        last_end_time = next_schedule.get('end_time')
                        j += 1
                    else:
                        break
            else:
                break

        # Create grouped entry
        first = group[0]
        last = group[-1]

        # Extract hour numbers from slot_index
        hours = []

        # Define standard school schedule times (07:00 start, 45 min per hour)
        # This maps start times to hour numbers
        time_to_hour_map = {
            '07:00': 1, '07:45': 2, '08:30': 3, '09:15': 4,
            '10:15': 5, '11:00': 6, '11:45': 7,  # After 1st break (09:15-10:15)
            '13:15': 8, '14:00': 9, '14:45': 10, '15:30': 11  # After lunch (11:45-13:15)
        }

        for s in group:
            # First try to use slot_index if it exists and seems valid
            if s.get('slot_index') is not None and s.get('slot_index') >= 0:
                hour_num = s['slot_index'] + 1  # slot_index is 0-based
                if hour_num not in hours:
                    hours.append(hour_num)
            # Otherwise try to derive from start_time using the map
            elif s.get('start_time') in time_to_hour_map:
                hour_num = time_to_hour_map[s['start_time']]
                if hour_num not in hours:
                    hours.append(hour_num)

        # If we still don't have valid hours or the count doesn't match, use fallback
        if not hours or len(hours) != len(group):
            # Fallback: use sequential numbering
            hours = list(range(1, len(group) + 1))
            logger.warning(f"Using fallback hour numbering for grouped schedule. Group size: {len(group)}, start_times: {[s.get('start_time') for s in group]}")

        # Sort hours to ensure proper order
        hours.sort()
        hour_range = "Jam ke-" + ", ".join(str(h) for h in hours)

        grouped_entry = {
            **first,  # Use first schedule as base
            'jtm_count': len(group),
            'hours': hours,
            'hour_range': hour_range,
            'schedule_ids': group_ids,
            'end_time': last.get('end_time'),  # Use last schedule's end time
            'time_range': f"{first.get('start_time')} - {last.get('end_time')}"
        }

        grouped.append(grouped_entry)
        i = j if j > i + 1 else i + 1

    return grouped


@router.post("/schedules")
async def create_schedule(payload: Dict, request: Request, user: Dict = Depends(get_current_user)):
    """Create schedule. Admin can create for any teacher. Guru/wali_kelas can create their OWN schedule (status: draft).

    Phase 6: blocks creation if guru/ruang bentrok (kecuali admin force=true).
    """
    is_admin = 'admin' in user.get('roles', [])
    teacher_id = payload.get('teacher_id')
    is_self_assign = teacher_id == user['id']
    can_self = bool(set(user.get('roles', [])) & {'guru', 'wali_kelas'})
    if not is_admin:
        if not (can_self and is_self_assign):
            raise HTTPException(403, "Hanya admin yang bisa menambahkan jadwal orang lain")

    # Conflict detection (Phase 6: guru + ruang)
    ay_id = payload.get('academic_year_id')
    if not ay_id:
        ay = await get_active_academic_year()
        ay_id = ay['id'] if ay else None
        payload['academic_year_id'] = ay_id
    force = payload.pop('force', False)
    if ay_id and payload.get('day') and payload.get('start_time') and payload.get('end_time'):
        conflicts = await _find_conflicts(
            ay_id, payload['day'], payload['start_time'], payload['end_time'],
            teacher_id=payload.get('teacher_id'),
            room_id=payload.get('room_id'),
            class_id=payload.get('class_id'),
        )
        blocking = conflicts['teacher'] + conflicts['room']
        if blocking and not (is_admin and force):
            raise HTTPException(status_code=409, detail={
                'message': f"Bentrok jadwal: {_conflict_message(conflicts)}",
                'conflicts': conflicts,
            })

    payload.setdefault('status', 'submitted' if is_admin else 'draft')
    payload['created_by'] = user['id']
    sched = ScheduleModel(**payload)
    doc = sched.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('submitted_at'):
        doc['submitted_at'] = doc['submitted_at'].isoformat() if isinstance(doc['submitted_at'], datetime) else doc['submitted_at']
    if doc.get('locked_at'):
        doc['locked_at'] = doc['locked_at'].isoformat() if isinstance(doc['locked_at'], datetime) else doc['locked_at']
    if is_admin and doc['status'] == 'submitted':
        doc['submitted_by'] = user['id']
        doc['submitted_at'] = now_wib().isoformat()
    await db.schedules.insert_one(doc)
    await log_audit(user, 'create', 'schedule', sched.id, details={'status': doc['status']}, request=request)
    return serialize_doc(doc)


# NOTE: /schedules/bulk-lock MUST be registered BEFORE /schedules/{sid}
# FastAPI matches in declaration order.
@router.put("/schedules/bulk-lock")
async def bulk_lock_schedules(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    """Lock multiple schedules at once."""
    ids = payload.get('ids') or []
    if not ids:
        raise HTTPException(400, "Tidak ada jadwal dipilih")
    res = await db.schedules.update_many({'id': {'$in': ids}}, {'$set': {
        'status': 'locked',
        'locked_at': now_wib().isoformat(),
        'locked_by': user['id'],
    }})
    await log_audit(user, 'bulk_lock', 'schedule', '-', details={'count': res.modified_count, 'requested': len(ids)}, request=request)
    return {'locked': res.modified_count, 'requested': len(ids)}


# my-today / grid / excel-template MUST come BEFORE /schedules/{sid} for path matching
@router.get("/schedules/my-today")
async def my_today_schedule(user: Dict = Depends(get_current_user)):
    ay = await get_active_academic_year()
    if not ay:
        return []
    day = current_day_id()
    items = await db.schedules.find({
        'teacher_id': user['id'], 'day': day, 'academic_year_id': ay['id'],
    }, {'_id': 0}).sort('start_time', 1).to_list(50)
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        journal = await db.journals.find_one({'schedule_id': s['id'], 'teacher_id': user['id']}, {'_id': 0, 'id': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['room_name'] = room.get('name') if room else None
        s['journal_filled'] = bool(journal)
        s['journal_id'] = journal.get('id') if journal else None
        enriched.append(serialize_doc(s))
    return enriched


@router.get("/schedules/grid")
async def schedules_grid(class_id: Optional[str] = None, teacher_id: Optional[str] = None,
                         user: Dict = Depends(get_current_user)):
    """Return schedule data structured as grid: days x slots"""
    settings = await get_settings()
    ay = await get_active_academic_year()
    if not ay:
        return {'days': [], 'slots': [], 'grid': {}}
    active_days = settings.get('active_days', ['senin','selasa','rabu','kamis','jumat'])
    # Default slots jika belum ada di settings
    default_slots = [
        {'name': 'Jam 1', 'start_time': '07:00', 'end_time': '07:45', 'is_break': False},
        {'name': 'Jam 2', 'start_time': '07:45', 'end_time': '08:30', 'is_break': False},
        {'name': 'Jam 3', 'start_time': '08:30', 'end_time': '09:15', 'is_break': False},
        {'name': 'Istirahat 1', 'start_time': '09:15', 'end_time': '09:30', 'is_break': True},
        {'name': 'Jam 4', 'start_time': '09:30', 'end_time': '10:15', 'is_break': False},
        {'name': 'Jam 5', 'start_time': '10:15', 'end_time': '11:00', 'is_break': False},
        {'name': 'Jam 6', 'start_time': '11:00', 'end_time': '11:45', 'is_break': False},
        {'name': 'Istirahat 2', 'start_time': '11:45', 'end_time': '12:15', 'is_break': True},
        {'name': 'Jam 7', 'start_time': '12:15', 'end_time': '13:00', 'is_break': False},
        {'name': 'Jam 8', 'start_time': '13:00', 'end_time': '13:45', 'is_break': False},
    ]
    slots = settings.get('teaching_slots', default_slots)
    q = {'academic_year_id': ay['id']}
    if class_id: q['class_id'] = class_id
    if teacher_id: q['teacher_id'] = teacher_id
    items = await db.schedules.find(q, {'_id': 0}).to_list(2000)
    enriched = []
    for s in items:
        cls = await db.classes.find_one({'id': s.get('class_id')}, {'_id': 0, 'name': 1})
        sub = await db.subjects.find_one({'id': s.get('subject_id')}, {'_id': 0, 'name': 1, 'code': 1})
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        room = await db.rooms.find_one({'id': s.get('room_id')}, {'_id': 0, 'name': 1})
        s['class_name'] = cls.get('name') if cls else None
        s['subject_name'] = sub.get('name') if sub else None
        s['subject_code'] = sub.get('code') if sub else None
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        s['room_name'] = room.get('name') if room else None
        enriched.append(serialize_doc(s))
    grid = {d: {} for d in active_days}
    for s in enriched:
        d = s['day']
        if d in grid:
            grid[d][s['start_time']] = s
    return {'days': active_days, 'slots': slots, 'grid': grid, 'schedules': enriched}


@router.get("/schedules/excel-template")
async def schedule_excel_template(user: Dict = Depends(require_role('admin'))):
    """Download Excel template for bulk schedule import"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal"
    headers = ['hari', 'jam_mulai', 'jam_selesai', 'kelas', 'mapel_kode', 'guru_username', 'ruang_kode', 'semester']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='006837')
        cell.alignment = Alignment(horizontal='center')
    examples = [
        ['senin', '07:00', '07:45', '7A', 'MTK', 'guru1', 'R-7A', 'ganjil'],
        ['senin', '07:45', '08:30', '7A', 'IPA', 'walas7a', 'R-7A', 'ganjil'],
        ['selasa', '07:00', '07:45', '7B', 'MTK', 'guru1', 'R-7B', 'ganjil'],
    ]
    for row in examples:
        ws.append(row)
    for col_letter, width in zip('ABCDEFGH', [10, 12, 12, 10, 12, 18, 12, 12]):
        ws.column_dimensions[col_letter].width = width

    ws2 = wb.create_sheet("INSTRUKSI")
    ws2.append(["Petunjuk Pengisian Template Jadwal"])
    ws2['A1'].font = Font(bold=True, size=14)
    instructions = [
        "",
        "1. Isi data jadwal pada sheet 'Jadwal' mulai baris 2.",
        "2. Kolom 'hari' wajib salah satu dari: senin, selasa, rabu, kamis, jumat, sabtu (huruf kecil).",
        "3. Kolom 'jam_mulai' dan 'jam_selesai' format HH:MM (24-jam), contoh: 07:00, 13:45.",
        "4. Kolom 'kelas' diisi NAMA kelas seperti yang terdaftar (contoh: 7A, 8B).",
        "5. Kolom 'mapel_kode' diisi KODE mapel (contoh: MTK, IPA, BIN).",
        "6. Kolom 'guru_username' diisi USERNAME guru pengampu.",
        "7. Kolom 'ruang_kode' diisi NAMA ruangan (contoh: R-7A).",
        "8. Kolom 'semester' diisi 'ganjil' atau 'genap' (untuk regular), atau '1','2','3','4','5','6' (untuk percepatan).",
        "9. Sistem akan mengabaikan baris kosong dan menolak baris dengan data tidak ditemukan.",
        "10. Setelah selesai, upload file melalui menu Admin > Jadwal > Import Excel.",
    ]
    for line in instructions:
        ws2.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="template_jadwal_matsandatama.xlsx"'},
    )


@router.post("/schedules/import-excel")
async def schedule_import_excel(file: UploadFile = File(...), request: Request = None,
                                user: Dict = Depends(require_role('admin'))):
    """Bulk import schedules from Excel file"""
    from openpyxl import load_workbook
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Hanya file .xlsx yang didukung")
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(400, "File terlalu besar (max 5MB)")
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb['Jadwal'] if 'Jadwal' in wb.sheetnames else wb.active
    except Exception as e:
        raise HTTPException(400, f"Gagal membaca Excel: {e}")

    ay = await get_active_academic_year()
    if not ay:
        raise HTTPException(400, "Tidak ada tahun pelajaran aktif")
    classes_map = {c['name']: c['id'] for c in await db.classes.find({'academic_year_id': ay['id']}, {'_id': 0}).to_list(500)}
    subjects_map = {s['code'].upper(): s['id'] for s in await db.subjects.find({}, {'_id': 0}).to_list(500)}
    rooms_map = {r['name']: r['id'] for r in await db.rooms.find({}, {'_id': 0}).to_list(500)}
    users_map = {u['username']: u['id'] for u in await db.users.find({}, {'_id': 0, 'username': 1, 'id': 1}).to_list(2000)}
    valid_days = {'senin', 'selasa', 'rabu', 'kamis', 'jumat', 'sabtu'}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success = 0
    errors = []
    new_docs = []
    for idx, row in enumerate(rows, start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        try:
            hari = str(row[0]).strip().lower() if row[0] else ''
            jm = str(row[1]).strip() if row[1] else ''
            js = str(row[2]).strip() if row[2] else ''
            kls = str(row[3]).strip() if row[3] else ''
            mp_kode = str(row[4]).strip().upper() if row[4] else ''
            gr_username = str(row[5]).strip() if row[5] else ''
            ruang = str(row[6]).strip() if row[6] else ''
            sem = str(row[7]).strip().lower() if row[7] else 'ganjil'

            if ':' not in jm:
                jm = str(jm)
            if ':' not in js:
                js = str(js)
            jm = jm[:5]
            js = js[:5]

            if hari not in valid_days:
                errors.append(f"Baris {idx}: hari '{hari}' tidak valid"); continue
            if kls not in classes_map:
                errors.append(f"Baris {idx}: kelas '{kls}' tidak ditemukan"); continue
            if mp_kode not in subjects_map:
                errors.append(f"Baris {idx}: mapel '{mp_kode}' tidak ditemukan"); continue
            if gr_username not in users_map:
                errors.append(f"Baris {idx}: guru '{gr_username}' tidak ditemukan"); continue
            if ruang not in rooms_map:
                errors.append(f"Baris {idx}: ruang '{ruang}' tidak ditemukan"); continue

            sched = {
                'id': str(uuid.uuid4()),
                'academic_year_id': ay['id'], 'semester': sem,
                'class_id': classes_map[kls], 'subject_id': subjects_map[mp_kode],
                'teacher_id': users_map[gr_username], 'room_id': rooms_map[ruang],
                'day': hari, 'start_time': jm, 'end_time': js,
                'is_published': True, 'created_at': datetime.utcnow().isoformat(),
            }
            new_docs.append(sched)
            success += 1
        except Exception as e:
            errors.append(f"Baris {idx}: {e}")
    if new_docs:
        await db.schedules.insert_many(new_docs)
    await log_audit(user, 'import_excel', 'schedule', None,
                    details={'success': success, 'errors': len(errors), 'filename': file.filename},
                    request=request)
    return {'success': success, 'errors': errors, 'total_rows': len(rows)}


# ============================================================
# WORKFLOW SUBMIT / LOCK / UNLOCK (also before /{sid})
# ============================================================
@router.put("/schedules/{sid}/submit")
async def submit_schedule(sid: str, request: Request, user: Dict = Depends(get_current_user)):
    """Guru/wali kelas kirim jadwal draft mereka ke admin untuk direview & dikunci."""
    existing = await db.schedules.find_one({'id': sid})
    if not existing:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_owner = existing.get('teacher_id') == user['id'] or existing.get('created_by') == user['id']
    if not (is_admin or is_owner):
        raise HTTPException(403, "Tidak diizinkan")
    if existing.get('status') == 'locked':
        raise HTTPException(400, "Jadwal sudah terkunci")
    await db.schedules.update_one({'id': sid}, {'$set': {
        'status': 'submitted',
        'submitted_at': now_wib().isoformat(),
        'submitted_by': user['id'],
    }})
    await log_audit(user, 'submit', 'schedule', sid, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@router.put("/schedules/{sid}/unsubmit")
async def unsubmit_schedule(sid: str, request: Request, user: Dict = Depends(get_current_user)):
    """Guru/wali kelas batal kirim jadwal yang sudah di-submit (kembali ke draft)."""
    existing = await db.schedules.find_one({'id': sid})
    if not existing:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_owner = existing.get('teacher_id') == user['id'] or existing.get('created_by') == user['id']
    if not (is_admin or is_owner):
        raise HTTPException(403, "Tidak diizinkan")
    if existing.get('status') not in ['submitted']:
        raise HTTPException(400, "Hanya jadwal dengan status 'terkirim' yang bisa dibatalkan")
    await db.schedules.update_one({'id': sid}, {'$set': {
        'status': 'draft',
        'submitted_at': None,
        'submitted_by': None,
    }})
    await log_audit(user, 'unsubmit', 'schedule', sid, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@router.put("/schedules/{sid}/approve")
async def approve_schedule(sid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    """Admin menyetujui jadwal yang sudah di-submit oleh guru."""
    print(f"[APPROVE] Mencari jadwal dengan id: {sid}")
    existing = await db.schedules.find_one({'id': sid})
    print(f"[APPROVE] Hasil pencarian: {existing is not None}")
    if existing:
        print(f"[APPROVE] Status jadwal: {existing.get('status')}")
    if not existing:
        # Coba cari dengan _id juga untuk debugging
        by_object_id = await db.schedules.find_one({'_id': sid})
        print(f"[APPROVE] Coba cari dengan _id: {by_object_id is not None}")
        raise HTTPException(404, "Tidak ditemukan")
    if existing.get('status') not in ['submitted', 'draft']:
        raise HTTPException(400, f"Jadwal dengan status '{existing.get('status')}' tidak bisa disetujui")
    await db.schedules.update_one({'id': sid}, {'$set': {
        'status': 'approved',
        'approved_at': now_wib().isoformat(),
        'approved_by': user['id'],
    }})
    await log_audit(user, 'approve', 'schedule', sid, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@router.put("/schedules/{sid}/lock")
async def lock_schedule(sid: str, request: Request, user: Dict = Depends(get_current_user)):
    """Lock jadwal. Admin bisa kunci semua. Wali kelas bisa kunci jadwal di kelasnya saja.

    Phase 6: tolak lock kalau ada konflik guru/ruang (kecuali admin force=true via query).
    """
    sch = await db.schedules.find_one({'id': sid})
    if not sch:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_wali = 'wali_kelas' in user.get('roles', [])
    can_lock = is_admin
    if is_wali and not is_admin:
        cls = await db.classes.find_one({'id': sch.get('class_id')}, {'_id': 0, 'homeroom_teacher_id': 1})
        can_lock = cls and cls.get('homeroom_teacher_id') == user['id']
    if not can_lock:
        raise HTTPException(403, "Tidak diizinkan mengunci jadwal ini")

    # Validasi: hanya bisa lock jadwal yang sudah approved
    if sch.get('status') not in ['approved', 'locked']:
        raise HTTPException(400, f"Jadwal harus disetujui terlebih dahulu sebelum dikunci. Status saat ini: {sch.get('status')}")

    # Conflict re-check (final guard before locking)
    force = request.query_params.get('force', '').lower() in ('true', '1', 'yes') if request else False
    conflicts = await _find_conflicts(
        sch.get('academic_year_id'), sch.get('day'), sch.get('start_time'), sch.get('end_time'),
        teacher_id=sch.get('teacher_id'), room_id=sch.get('room_id'),
        class_id=sch.get('class_id'), exclude_id=sid,
    )
    blocking = conflicts['teacher'] + conflicts['room']
    if blocking and not (is_admin and force):
        raise HTTPException(status_code=409, detail={
            'message': f"Tidak bisa lock: {_conflict_message(conflicts)}",
            'conflicts': conflicts,
        })

    locked_by_role = 'admin' if is_admin else 'wali_kelas'
    await db.schedules.update_one({'id': sid}, {'$set': {
        'status': 'locked',
        'locked_at': now_wib().isoformat(),
        'locked_by': user['id'],
        'locked_by_role': locked_by_role,
    }})
    await log_audit(user, 'lock', 'schedule', sid, details={'by_role': locked_by_role}, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@router.put("/schedules/{sid}/unlock")
async def unlock_schedule(sid: str, request: Request, user: Dict = Depends(get_current_user)):
    """Unlock. Admin bisa unlock kapan saja. Wali kelas HANYA bisa unlock jadwal yang dia sendiri kunci.
    Jika jadwal dikunci admin, wali kelas TIDAK BISA membukanya."""
    sch = await db.schedules.find_one({'id': sid})
    if not sch:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    locked_by_role = sch.get('locked_by_role', 'admin')
    locked_by_uid = sch.get('locked_by')
    if not is_admin:
        if locked_by_role != 'wali_kelas' or locked_by_uid != user['id']:
            raise HTTPException(403, "Hanya admin yang bisa membuka kunci jadwal ini (dikunci admin)")
    await db.schedules.update_one({'id': sid}, {'$set': {
        'status': 'submitted',
        'locked_at': None,
        'locked_by': None,
        'locked_by_role': None,
    }})
    await log_audit(user, 'unlock', 'schedule', sid, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


# ============================================================
# UPDATE / DELETE by ID (generic params LAST)
# ============================================================
@router.put("/schedules/{sid}")
async def update_schedule(sid: str, payload: Dict, request: Request, user: Dict = Depends(get_current_user)):
    existing = await db.schedules.find_one({'id': sid})
    if not existing:
        raise HTTPException(404, "Jadwal tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_owner = existing.get('teacher_id') == user['id'] or existing.get('created_by') == user['id']
    status_val = existing.get('status', 'submitted')
    if status_val == 'locked' and not is_admin:
        raise HTTPException(403, "Jadwal sudah dikunci. Hanya admin yang bisa edit setelah dibuka kuncinya.")
    if status_val == 'submitted' and not is_admin:
        raise HTTPException(403, "Jadwal sudah dikirim. Hanya admin yang bisa edit.")
    if not (is_admin or (is_owner and status_val == 'draft')):
        raise HTTPException(403, "Tidak diizinkan")
    payload.pop('status', None)
    payload.pop('submitted_at', None); payload.pop('submitted_by', None)
    payload.pop('locked_at', None); payload.pop('locked_by', None)
    payload.pop('id', None); payload.pop('_id', None); payload.pop('created_by', None)
    await db.schedules.update_one({'id': sid}, {'$set': payload})
    await log_audit(user, 'update', 'schedule', sid, request=request)
    doc = await db.schedules.find_one({'id': sid}, {'_id': 0})
    return serialize_doc(doc)


@router.delete("/schedules/{sid}")
async def delete_schedule(sid: str, request: Request, user: Dict = Depends(get_current_user)):
    existing = await db.schedules.find_one({'id': sid})
    if not existing:
        raise HTTPException(404, "Tidak ditemukan")
    is_admin = 'admin' in user.get('roles', [])
    is_owner = existing.get('teacher_id') == user['id'] or existing.get('created_by') == user['id']
    if existing.get('status') == 'locked' and not is_admin:
        raise HTTPException(403, "Jadwal terkunci tidak bisa dihapus")
    if not (is_admin or (is_owner and existing.get('status', 'submitted') == 'draft')):
        raise HTTPException(403, "Tidak diizinkan")
    await db.schedules.delete_one({'id': sid})
    await log_audit(user, 'delete', 'schedule', sid, request=request)
    return {'message': 'Dihapus'}


# ============================================================
# PIKET SCHEDULES (Jadwal Piket)
# ============================================================
@router.get("/piket-schedules")
async def list_piket(day: Optional[str] = None, user: Dict = Depends(get_current_user)):
    q = {}
    if day:
        q['day'] = day
    items = await db.piket_schedules.find(q, {'_id': 0}).sort([('day', 1), ('start_time', 1)]).to_list(500)
    enriched = []
    for s in items:
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        enriched.append(serialize_doc(s))
    return enriched


@router.post("/piket-schedules")
async def create_piket(payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    doc = {
        'id': str(uuid.uuid4()),
        'day': payload.get('day', 'senin'),
        'shift': payload.get('shift', 'pagi'),
        'start_time': payload.get('start_time', '06:30'),
        'end_time': payload.get('end_time', '14:00'),
        'teacher_id': payload.get('teacher_id'),
        'notes': payload.get('notes', ''),
        'is_active': payload.get('is_active', True),
        'created_at': datetime.utcnow().isoformat(),
    }
    await db.piket_schedules.insert_one(doc)
    await log_audit(user, 'create', 'piket_schedule', doc['id'],
                    details={'day': doc['day'], 'teacher_id': doc['teacher_id']}, request=request)
    return serialize_doc(doc)


# specific BEFORE /{pid}
@router.get("/piket-schedules/today")
async def piket_today(user: Dict = Depends(get_current_user)):
    day = current_day_id()
    items = await db.piket_schedules.find({'day': day, 'is_active': True}, {'_id': 0}).sort('start_time', 1).to_list(50)
    enriched = []
    for s in items:
        teacher = await db.users.find_one({'id': s.get('teacher_id')}, {'_id': 0, 'full_name': 1})
        s['teacher_name'] = teacher.get('full_name') if teacher else None
        enriched.append(serialize_doc(s))
    return enriched


@router.put("/piket-schedules/{pid}")
async def update_piket(pid: str, payload: Dict, request: Request, user: Dict = Depends(require_role('admin'))):
    payload.pop('_id', None)
    payload.pop('id', None)
    res = await db.piket_schedules.update_one({'id': pid}, {'$set': payload})
    if res.matched_count == 0:
        raise HTTPException(404, "Jadwal piket tidak ditemukan")
    await log_audit(user, 'update', 'piket_schedule', pid, details=payload, request=request)
    doc = await db.piket_schedules.find_one({'id': pid}, {'_id': 0})
    return serialize_doc(doc)


@router.delete("/piket-schedules/{pid}")
async def delete_piket(pid: str, request: Request, user: Dict = Depends(require_role('admin'))):
    await db.piket_schedules.delete_one({'id': pid})
    await log_audit(user, 'delete', 'piket_schedule', pid, request=request)
    return {'message': 'Dihapus'}

