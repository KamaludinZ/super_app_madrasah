"""
Router untuk RKAM (Rencana Kegiatan dan Anggaran Madrasah).
"""
from datetime import datetime
from typing import Dict, List, Optional
import uuid

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient

from core import db, require_role, get_current_user, logger, serialize_doc
from models import RKAMBudgetItemModel, RKAMDocumentModel
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(tags=["RKAM"])


# ============================================================================
# BUDGET ITEMS ENDPOINTS
# ============================================================================

@router.get("/rkam/budget-items")
async def list_budget_items(
    fiscal_year: Optional[str] = None,
    quarter: Optional[str] = None,
    category: Optional[str] = None,
    bidang: Optional[str] = None,
    sumber_dana: Optional[str] = None,
    is_active: Optional[bool] = True,
    user: Dict = Depends(get_current_user)
):
    """
    List RKAM budget items.

    Public endpoint - anyone authenticated can view.
    Filters by fiscal_year, quarter, category, bidang, sumber_dana, is_active.
    """
    try:
        logger.info(f"[RKAM GET] List budget items - fiscal_year={fiscal_year}, user={user.get('username')}")

        query = {}
        if fiscal_year:
            query['fiscal_year'] = fiscal_year
        if quarter:
            query['quarter'] = quarter
        if category:
            query['category'] = category
        if bidang:
            query['bidang'] = bidang
        if sumber_dana:
            query['sumber_dana'] = sumber_dana
        if is_active is not None:
            query['is_active'] = is_active

        logger.info(f"[RKAM GET] Query: {query}")

        items = await db.rkam_budget_items.find(query).sort('code', 1).to_list(None)
        logger.info(f"[RKAM GET] Found {len(items)} items")

        # Serialize and enrich with creator names
        serialized_items = []
        for idx, item in enumerate(items):
            try:
                # Serialize to remove MongoDB _id
                item_clean = serialize_doc(item)

                # Enrich with creator name
                if item_clean.get('created_by'):
                    creator = await db.users.find_one({'id': item_clean['created_by']})
                    if creator:
                        item_clean['created_by_name'] = creator.get('full_name', 'Unknown')

                serialized_items.append(item_clean)
            except Exception as e:
                logger.error(f"[RKAM GET] Error processing item {idx}: {e}")

        logger.info(f"[RKAM GET] Returning {len(serialized_items)} items")
        return serialized_items
    except Exception as e:
        logger.error(f"[RKAM GET] ERROR: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to list budget items: {str(e)}")


@router.post("/rkam/budget-items")
async def create_budget_item(
    payload: Dict,
    request: Request,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Create new RKAM budget item.

    Only Admin and Bendahara can create.
    """
    try:
        logger.info(f"[RKAM POST] Received payload: {payload}")
        logger.info(f"[RKAM POST] User: {user.get('id')} ({user.get('username')})")

        # Validate required fields
        if not payload.get('name'):
            raise HTTPException(status_code=400, detail="Field 'name' is required")
        if not payload.get('fiscal_year'):
            raise HTTPException(status_code=400, detail="Field 'fiscal_year' is required")

        # v1.1.1: Support dual budget fields (BOS & Komite)
        allocated_bos = float(payload.get('allocated_bos', 0.0))
        allocated_komite = float(payload.get('allocated_komite', 0.0))
        realized_bos = float(payload.get('realized_bos', 0.0))
        realized_komite = float(payload.get('realized_komite', 0.0))

        item = RKAMBudgetItemModel(
            code=payload.get('code'),
            name=payload['name'],
            category=payload.get('category'),
            bidang=payload.get('bidang'),
            description=payload.get('description'),
            # v1.1.1: Dual budget fields
            allocated_bos=allocated_bos,
            allocated_komite=allocated_komite,
            realized_bos=realized_bos,
            realized_komite=realized_komite,
            # Deprecated fields (kept for backward compatibility)
            allocated_amount=allocated_bos + allocated_komite,
            realized_amount=realized_bos + realized_komite,
            remaining_amount=(allocated_bos + allocated_komite) - (realized_bos + realized_komite),
            sumber_dana=payload.get('sumber_dana'),
            fiscal_year=payload['fiscal_year'],
            quarter=payload.get('quarter'),
            month=payload.get('month'),
            created_by=user['id']
        )

        logger.info(f"[RKAM POST] Created model: id={item.id}, name={item.name}")

        item_dict = item.model_dump()
        logger.info(f"[RKAM POST] Inserting to database: {item_dict}")

        result = await db.rkam_budget_items.insert_one(item_dict)

        logger.info(f"[RKAM POST] Insert result: inserted_id={result.inserted_id}, acknowledged={result.acknowledged}")

        if not result.inserted_id:
            raise HTTPException(status_code=500, detail="Failed to create budget item")

        # Verify data was saved
        saved_item = await db.rkam_budget_items.find_one({'id': item.id})
        logger.info(f"[RKAM POST] Verification - Item saved: {saved_item is not None}")

        return {"id": item.id, "message": "Budget item created successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[RKAM POST] ERROR: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create budget item: {str(e)}")


@router.put("/rkam/budget-items/{item_id}")
async def update_budget_item(
    item_id: str,
    payload: Dict,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Update RKAM budget item.

    Only Admin and Bendahara can update.
    """
    existing = await db.rkam_budget_items.find_one({'id': item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Budget item not found")

    # v1.1.1: Support dual budget fields (BOS & Komite)
    allocated_bos = float(payload.get('allocated_bos', existing.get('allocated_bos', 0.0)))
    allocated_komite = float(payload.get('allocated_komite', existing.get('allocated_komite', 0.0)))
    realized_bos = float(payload.get('realized_bos', existing.get('realized_bos', 0.0)))
    realized_komite = float(payload.get('realized_komite', existing.get('realized_komite', 0.0)))

    update_data = {
        'code': payload.get('code', existing.get('code')),
        'name': payload.get('name', existing.get('name')),
        'category': payload.get('category', existing.get('category')),
        'bidang': payload.get('bidang', existing.get('bidang')),
        'description': payload.get('description', existing.get('description')),
        # v1.1.1: Dual budget fields
        'allocated_bos': allocated_bos,
        'allocated_komite': allocated_komite,
        'realized_bos': realized_bos,
        'realized_komite': realized_komite,
        # Deprecated fields (kept for backward compatibility)
        'allocated_amount': allocated_bos + allocated_komite,
        'realized_amount': realized_bos + realized_komite,
        'remaining_amount': (allocated_bos + allocated_komite) - (realized_bos + realized_komite),
        'sumber_dana': payload.get('sumber_dana', existing.get('sumber_dana')),
        'fiscal_year': payload.get('fiscal_year', existing.get('fiscal_year')),
        'quarter': payload.get('quarter', existing.get('quarter')),
        'month': payload.get('month', existing.get('month')),
        'is_active': payload.get('is_active', existing.get('is_active', True)),
        'updated_at': datetime.utcnow(),
        'updated_by': user['id']
    }

    await db.rkam_budget_items.update_one({'id': item_id}, {'$set': update_data})

    return {"message": "Budget item updated successfully"}


@router.delete("/rkam/budget-items/{item_id}")
async def delete_budget_item(
    item_id: str,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Delete RKAM budget item (soft delete).

    Only Admin and Bendahara can delete.
    """
    result = await db.rkam_budget_items.update_one(
        {'id': item_id},
        {'$set': {'is_active': False, 'updated_at': datetime.utcnow(), 'updated_by': user['id']}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Budget item not found")

    return {"message": "Budget item deleted successfully"}


# ============================================================================
# IMPORT / EXPORT ENDPOINTS
# ============================================================================

@router.get("/rkam/budget-items/export")
async def export_budget_items(
    fiscal_year: Optional[str] = None,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Export RKAM budget items to Excel.

    Only Admin and Bendahara can export.
    """
    query = {'is_active': True}
    if fiscal_year:
        query['fiscal_year'] = fiscal_year

    items = await db.rkam_budget_items.find(query).sort('code', 1).to_list(None)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RKAM Budget"

    # Header styling
    header_fill = PatternFill(start_color="006837", end_color="006837", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Kode", "Nama Item", "Kategori", "Deskripsi",
        "Anggaran Dialokasikan", "Realisasi", "Sisa",
        "Tahun Anggaran", "Triwulan", "Bulan"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.get('code', '')).border = border
        ws.cell(row=row_num, column=2, value=item.get('name', '')).border = border
        ws.cell(row=row_num, column=3, value=item.get('category', '')).border = border
        ws.cell(row=row_num, column=4, value=item.get('description', '')).border = border
        ws.cell(row=row_num, column=5, value=item.get('allocated_amount', 0.0)).border = border
        ws.cell(row=row_num, column=6, value=item.get('realized_amount', 0.0)).border = border
        ws.cell(row=row_num, column=7, value=item.get('remaining_amount', 0.0)).border = border
        ws.cell(row=row_num, column=8, value=item.get('fiscal_year', '')).border = border
        ws.cell(row=row_num, column=9, value=item.get('quarter', '')).border = border
        ws.cell(row=row_num, column=10, value=item.get('month', '')).border = border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"RKAM_Budget_{fiscal_year or 'All'}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/rkam/budget-items/import")
async def import_budget_items(
    file: UploadFile = File(...),
    fiscal_year: str = None,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Import RKAM budget items from Excel.

    Only Admin and Bendahara can import.
    Expected columns: Kode, Nama Item, Kategori, Deskripsi,
                      Anggaran Dialokasikan, Realisasi, Sisa,
                      Tahun Anggaran, Triwulan, Bulan
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    imported_count = 0
    errors = []

    # Skip header row
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not row[1]:  # Skip if name is empty
                continue

            item = RKAMBudgetItemModel(
                code=str(row[0]) if row[0] else None,
                name=str(row[1]),
                category=str(row[2]) if row[2] else None,
                description=str(row[3]) if row[3] else None,
                allocated_amount=float(row[4]) if row[4] else 0.0,
                realized_amount=float(row[5]) if row[5] else 0.0,
                remaining_amount=float(row[6]) if row[6] else 0.0,
                fiscal_year=str(row[7]) if row[7] else (fiscal_year or ''),
                quarter=str(row[8]) if row[8] else None,
                month=int(row[9]) if row[9] else None,
                created_by=user['id']
            )

            await db.rkam_budget_items.insert_one(item.model_dump())
            imported_count += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    return {
        "message": f"Import completed. {imported_count} items imported.",
        "imported_count": imported_count,
        "errors": errors if errors else None
    }


# ============================================================================
# DOCUMENTS ENDPOINTS
# ============================================================================

@router.get("/rkam/documents")
async def list_documents(
    fiscal_year: Optional[str] = None,
    is_public: Optional[bool] = None,
    document_type: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    """
    List RKAM documents.

    All authenticated users can view public documents.
    Only Admin and Bendahara can view all documents.
    """
    try:
        logger.info(f"[RKAM DOC GET] List documents - fiscal_year={fiscal_year}, is_public={is_public}, document_type={document_type}, user={user.get('username')}")

        query = {'is_active': True}

        # Admin and Bendahara can see all documents
        is_admin_or_bendahara = any(role in user.get('roles', []) for role in ['admin', 'bendahara'])

        if not is_admin_or_bendahara:
            # Non-admin can only see public documents
            query['is_public'] = True
        elif is_public is not None:
            query['is_public'] = is_public

        if fiscal_year:
            query['fiscal_year'] = fiscal_year
        if document_type:
            query['document_type'] = document_type

        logger.info(f"[RKAM DOC GET] Query: {query}")

        docs = await db.rkam_documents.find(query).sort('upload_date', -1).to_list(None)
        logger.info(f"[RKAM DOC GET] Found {len(docs)} documents")

        # Serialize and enrich with uploader names
        serialized_docs = []
        for idx, doc in enumerate(docs):
            try:
                doc_clean = serialize_doc(doc)  # Remove MongoDB _id field
                if doc_clean.get('uploaded_by'):
                    uploader = await db.users.find_one({'id': doc_clean['uploaded_by']})
                    if uploader:
                        doc_clean['uploaded_by_name'] = uploader.get('full_name', 'Unknown')
                serialized_docs.append(doc_clean)
            except Exception as e:
                logger.error(f"[RKAM DOC GET] Error processing document {idx}: {e}")

        logger.info(f"[RKAM DOC GET] Returning {len(serialized_docs)} serialized documents")
        return serialized_docs
    except Exception as e:
        logger.error(f"[RKAM DOC GET] ERROR: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to list documents: {str(e)}")


@router.get("/public/rkam/documents")
async def list_public_documents(
    fiscal_year: Optional[str] = None
):
    """
    Public endpoint for RKAM documents.

    No authentication required.
    Only returns public documents.
    """
    try:
        logger.info(f"[RKAM DOC PUBLIC] List public documents - fiscal_year={fiscal_year}")

        query = {'is_active': True, 'is_public': True}

        if fiscal_year:
            query['fiscal_year'] = fiscal_year

        logger.info(f"[RKAM DOC PUBLIC] Query: {query}")

        docs = await db.rkam_documents.find(query).sort('upload_date', -1).to_list(None)
        logger.info(f"[RKAM DOC PUBLIC] Found {len(docs)} documents")

        # Serialize documents
        serialized_docs = [serialize_doc(doc) for doc in docs]

        return serialized_docs
    except Exception as e:
        logger.error(f"[RKAM DOC PUBLIC] ERROR: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get public documents: {str(e)}")


@router.post("/rkam/documents")
async def create_document(
    payload: Dict,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Create new RKAM document entry.

    Only Admin and Bendahara can create.
    Note: File upload should be handled separately.
    """
    try:
        logger.info(f"[RKAM DOC POST] Received payload: {payload}")
        logger.info(f"[RKAM DOC POST] User: {user.get('id')} ({user.get('username')})")

        # Validate required fields
        if not payload.get('title'):
            raise HTTPException(status_code=400, detail="Field 'title' is required")
        if not payload.get('document_type'):
            raise HTTPException(status_code=400, detail="Field 'document_type' is required")
        if not payload.get('file_url'):
            raise HTTPException(status_code=400, detail="Field 'file_url' is required")
        if not payload.get('file_name'):
            raise HTTPException(status_code=400, detail="Field 'file_name' is required")
        if not payload.get('fiscal_year'):
            raise HTTPException(status_code=400, detail="Field 'fiscal_year' is required")

        doc = RKAMDocumentModel(
            title=payload['title'],
            description=payload.get('description'),
            document_type=payload['document_type'],
            file_url=payload['file_url'],
            file_name=payload['file_name'],
            file_size=payload.get('file_size'),
            mime_type=payload.get('mime_type'),
            fiscal_year=payload['fiscal_year'],
            quarter=payload.get('quarter'),
            is_public=payload.get('is_public', True),
            uploaded_by=user['id'],
            tags=payload.get('tags', [])
        )

        logger.info(f"[RKAM DOC POST] Created model: id={doc.id}, title={doc.title}")

        doc_dict = doc.model_dump()
        logger.info(f"[RKAM DOC POST] Inserting to database")

        result = await db.rkam_documents.insert_one(doc_dict)

        logger.info(f"[RKAM DOC POST] Insert result: inserted_id={result.inserted_id}, acknowledged={result.acknowledged}")

        if not result.inserted_id:
            raise HTTPException(status_code=500, detail="Failed to create document")

        # Verify data was saved
        saved_doc = await db.rkam_documents.find_one({'id': doc.id})
        logger.info(f"[RKAM DOC POST] Verification - Document saved: {saved_doc is not None}")

        return {"id": doc.id, "message": "Document created successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[RKAM DOC POST] ERROR: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create document: {str(e)}")


@router.put("/rkam/documents/{doc_id}")
async def update_document(
    doc_id: str,
    payload: Dict,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Update RKAM document metadata.

    Only Admin and Bendahara can update.
    """
    existing = await db.rkam_documents.find_one({'id': doc_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Document not found")

    update_data = {
        'title': payload.get('title', existing.get('title')),
        'description': payload.get('description', existing.get('description')),
        'document_type': payload.get('document_type', existing.get('document_type')),
        'fiscal_year': payload.get('fiscal_year', existing.get('fiscal_year')),
        'quarter': payload.get('quarter', existing.get('quarter')),
        'is_public': payload.get('is_public', existing.get('is_public', True)),
        'tags': payload.get('tags', existing.get('tags', []))
    }

    await db.rkam_documents.update_one({'id': doc_id}, {'$set': update_data})

    return {"message": "Document updated successfully"}


@router.delete("/rkam/documents/{doc_id}")
async def delete_document(
    doc_id: str,
    user: Dict = Depends(require_role('admin', 'bendahara'))
):
    """
    Delete RKAM document (soft delete).

    Only Admin and Bendahara can delete.
    """
    result = await db.rkam_documents.update_one(
        {'id': doc_id},
        {'$set': {'is_active': False}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")

    return {"message": "Document deleted successfully"}


# ============================================================================
# PUBLIC ENDPOINTS
# ============================================================================

@router.get("/public/rkam/budget-items")
async def get_public_budget_items(
    fiscal_year: Optional[str] = None,
    bidang: Optional[str] = None,
    page: int = 1,
    limit: int = 20
):
    """
    Public endpoint for RKAM budget items list.

    No authentication required.
    Returns list of budget items with pagination.
    """
    try:
        query = {'is_active': True}
        if fiscal_year:
            query['fiscal_year'] = fiscal_year
        if bidang:
            query['bidang'] = bidang

        # Count total items
        total = await db.rkam_budget_items.count_documents(query)

        # Calculate skip
        skip = (page - 1) * limit

        # Get items with pagination
        items = await db.rkam_budget_items.find(query).sort('code', 1).skip(skip).limit(limit).to_list(None)

        # Serialize items
        serialized_items = [serialize_doc(item) for item in items]

        return {
            'items': serialized_items,
            'total': total,
            'page': page,
            'limit': limit,
            'total_pages': (total + limit - 1) // limit  # Ceiling division
        }
    except Exception as e:
        logger.error(f"[RKAM PUBLIC] ERROR: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get budget items: {str(e)}")


@router.get("/public/rkam/budget-summary")
async def get_public_budget_summary(
    fiscal_year: Optional[str] = None,
    bidang: Optional[str] = None
):
    """
    Public endpoint for RKAM budget summary.

    No authentication required.
    Returns aggregated budget data by category and bidang.
    """
    try:
        logger.info(f"[RKAM] GET budget-summary: fiscal_year={fiscal_year}, bidang={bidang}")

        query = {'is_active': True}
        if fiscal_year:
            query['fiscal_year'] = fiscal_year
        if bidang:
            query['bidang'] = bidang

        logger.info(f"[RKAM] Query: {query}")
        items = await db.rkam_budget_items.find(query).to_list(None)
        logger.info(f"[RKAM] Found {len(items)} items")

        # Calculate totals
        total_allocated = sum(item.get('allocated_amount', 0.0) for item in items)
        total_realized = sum(item.get('realized_amount', 0.0) for item in items)
        total_remaining = sum(item.get('remaining_amount', 0.0) for item in items)

        # Group by sumber_dana
        sumber_dana_groups = {}
        for item in items:
            sumber = item.get('sumber_dana', 'LAINNYA')
            if sumber not in sumber_dana_groups:
                sumber_dana_groups[sumber] = {
                    'sumber_dana': sumber,
                    'allocated': 0.0,
                    'realized': 0.0,
                    'remaining': 0.0
                }
            sumber_dana_groups[sumber]['allocated'] += item.get('allocated_amount', 0.0)
            sumber_dana_groups[sumber]['realized'] += item.get('realized_amount', 0.0)
            sumber_dana_groups[sumber]['remaining'] += item.get('remaining_amount', 0.0)

        # Group by category
        categories = {}
        for item in items:
            cat = item.get('category', 'Lainnya')
            if cat not in categories:
                categories[cat] = {
                    'category': cat,
                    'allocated': 0.0,
                    'realized': 0.0,
                    'remaining': 0.0,
                    'items': []
                }
            categories[cat]['allocated'] += item.get('allocated_amount', 0.0)
            categories[cat]['realized'] += item.get('realized_amount', 0.0)
            categories[cat]['remaining'] += item.get('remaining_amount', 0.0)
            categories[cat]['items'].append({
                'id': item.get('id'),
                'code': item.get('code'),
                'name': item.get('name'),
                'sumber_dana': item.get('sumber_dana'),
                'allocated': item.get('allocated_amount', 0.0),
                'realized': item.get('realized_amount', 0.0),
                'remaining': item.get('remaining_amount', 0.0)
            })

        logger.info(f"[RKAM] Response prepared successfully")
        return {
            'fiscal_year': fiscal_year,
            'total_allocated': total_allocated,
            'total_realized': total_realized,
            'total_remaining': total_remaining,
            'percentage_realized': (total_realized / total_allocated * 100) if total_allocated > 0 else 0,
            'sumber_dana_groups': list(sumber_dana_groups.values()),
            'categories': list(categories.values())
        }
    except Exception as e:
        logger.error(f"[RKAM] ERROR in budget-summary: {str(e)}")
        logger.error(f"[RKAM] Error details:", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
