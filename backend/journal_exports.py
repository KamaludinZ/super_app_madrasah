import calendar
import io
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BRAND_HEX = "006837"
GOLD_HEX = "C8A24A"

DAY_ID_TO_LABEL = {
    "senin": "Senin",
    "selasa": "Selasa",
    "rabu": "Rabu",
    "kamis": "Kamis",
    "jumat": "Jumat",
    "sabtu": "Sabtu",
    "minggu": "Minggu",
}

DAY_IDX_TO_ID = {
    0: "senin",
    1: "selasa",
    2: "rabu",
    3: "kamis",
    4: "jumat",
    5: "sabtu",
    6: "minggu",
}


def _safe_date_label(iso_dt: str) -> Tuple[str, str]:
    try:
        dt = datetime.fromisoformat(iso_dt.replace("Z", "+00:00"))
        day_id = DAY_IDX_TO_ID.get(dt.weekday(), "")
        day_label = DAY_ID_TO_LABEL.get(day_id, "")
        date_label = dt.strftime("%d-%m-%Y")
        return day_label, date_label
    except Exception:
        return "-", "-"


def _format_absensi_text(journal: Dict[str, Any], student_name_map: Dict[str, str]) -> str:
    details = journal.get("attendance_details") or []
    if not details:
        return (
            f"Hadir:{journal.get('siswa_hadir', 0)} | "
            f"Sakit:{journal.get('siswa_sakit', 0)} | "
            f"Izin:{journal.get('siswa_izin', 0)} | "
            f"Alpa:{journal.get('siswa_tidak_hadir', 0)}"
        )

    absent_lines: List[str] = []
    for item in details:
        status = (item.get("status") or "").lower()
        if status in ("alpa", "izin", "sakit"):
            sid = item.get("student_id")
            sname = student_name_map.get(sid, sid or "-")
            ket = item.get("note") or "-"
            absent_lines.append(f"{sname} ({status.upper()}; {ket})")

    if absent_lines:
        return (
            f"Jml Tidak Hadir: {len(absent_lines)}\n"
            + "; ".join(absent_lines)
        )

    return "Semua hadir"


def _header_style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BRAND_HEX)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thick", color=GOLD_HEX))
    ws.row_dimensions[1].height = 28


def export_monthly_teacher_journal_excel(
    journals: List[Dict[str, Any]],
    teacher_name: str,
    month: int,
    year: int,
    student_name_map: Dict[str, str],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jurnal Bulanan"

    headers = [
        "NO",
        "HARI, TANGGAL",
        "JAM KE",
        "KELAS",
        "KD/INDIKATOR",
        "MATERI/POKOK BAHASAN",
        "ABSENSI SISWA",
    ]
    ws.append(headers)
    _header_style(ws)

    for idx, j in enumerate(journals, start=1):
        day_label, date_label = _safe_date_label(j.get("started_at", ""))
        jam_ke = j.get("slot_index")
        jam_ke_label = f"Ke-{jam_ke + 1}" if isinstance(jam_ke, int) else f"{j.get('scheduled_start', '-')}-{j.get('scheduled_end', '-')}"
        kd = j.get("kd_indikator") or "-"
        materi = j.get("materi") or "-"
        absensi = _format_absensi_text(j, student_name_map)

        ws.append([
            idx,
            f"{day_label}, {date_label}",
            jam_ke_label,
            j.get("class_name") or "-",
            kd,
            materi,
            absensi,
        ])

    widths = [6, 22, 14, 12, 20, 34, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 34
        for c in range(1, 8):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def export_monthly_teacher_journal_pdf(
    journals: List[Dict[str, Any]],
    teacher_name: str,
    teacher_nip: str,
    month: int,
    year: int,
    city_label: str,
    head_name: str,
    head_nip: str,
    student_name_map: Dict[str, str],
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    story = []

    title = f"LAPORAN JURNAL HARIAN GURU - {teacher_name} ({month:02d}/{year})"
    story.append(Paragraph(title, styles["Heading3"]))
    story.append(Spacer(1, 8))

    table_data = [[
        "NO",
        "HARI, TANGGAL",
        "JAM KE",
        "KELAS",
        "KD/INDIKATOR",
        "MATERI/POKOK BAHASAN",
        "ABSENSI SISWA",
    ]]

    for idx, j in enumerate(journals, start=1):
        day_label, date_label = _safe_date_label(j.get("started_at", ""))
        jam_ke = j.get("slot_index")
        jam_ke_label = f"Ke-{jam_ke + 1}" if isinstance(jam_ke, int) else f"{j.get('scheduled_start', '-')}-{j.get('scheduled_end', '-')}"
        kd = j.get("kd_indikator") or "-"
        materi = j.get("materi") or "-"
        absensi = _format_absensi_text(j, student_name_map)
        table_data.append([
            str(idx),
            f"{day_label}, {date_label}",
            jam_ke_label,
            j.get("class_name") or "-",
            kd,
            materi,
            absensi,
        ])

    col_widths = [28, 95, 62, 58, 86, 150, 220]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#006837")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 16))

    today_str = datetime.now().strftime("%d-%m-%Y")
    sign_data = [
        ["", f"{city_label.upper()}, {today_str}"],
        ["MENGETAHUI", ""],
        ["KEPALA MTSN2", "GURU ..."],
        ["", ""],
        ["", ""],
        [head_name or "-", teacher_name or "-"],
        [head_nip or "-", teacher_nip or "-"],
    ]
    sign_tbl = Table(sign_data, colWidths=[390, 390])
    sign_tbl.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),
    ]))
    story.append(sign_tbl)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf
