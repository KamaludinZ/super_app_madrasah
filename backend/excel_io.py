"""
Excel import/export helpers for Super Apps MATSANDATAMA.
Reusable functions for generating templates and parsing uploads.
"""
import io
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BRAND_HEX = "006837"
GOLD_HEX = "C8A24A"


def _make_workbook_with_data(sheet_name: str, headers: List[str],
                              examples: List[List[Any]], col_widths: List[int],
                              instructions: List[str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=BRAND_HEX)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thick', color=GOLD_HEX))
    ws.row_dimensions[1].height = 26
    # Example rows
    for row in examples:
        ws.append(row)
    # Column widths
    for idx, w in enumerate(col_widths):
        letter = ws.cell(row=1, column=idx + 1).column_letter
        ws.column_dimensions[letter].width = w

    # INSTRUKSI sheet
    ws2 = wb.create_sheet("INSTRUKSI")
    ws2['A1'] = "Petunjuk Pengisian Template"
    ws2['A1'].font = Font(bold=True, size=14, color=BRAND_HEX)
    ws2.column_dimensions['A'].width = 100
    for line in [''] + instructions:
        ws2.append([line])
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# USER TEMPLATE
# ============================================================
def user_template() -> bytes:
    wb = _make_workbook_with_data(
        sheet_name="Pengguna",
        headers=['username', 'password', 'nama_lengkap', 'roles', 'nip_nuptk', 'nisn',
                 'email', 'phone', 'gender', 'kelas_siswa', 'wali_kelas'],
        examples=[
            ['guru_baru1', 'pwd123', 'Drs. Andi Pranowo', 'guru', '198001012005011001', '', 'andi@matsa.sch.id', '081234567890', 'L', '', ''],
            ['siswa_baru1', 'pwd123', 'Bagus Pratama', 'siswa', '', '0098765499', '', '', 'L', '7A', ''],
            ['walas_baru', 'pwd123', 'Ibu Rina Sari, S.Pd', 'wali_kelas,guru', '199105072015012001', '', 'rina@matsa.sch.id', '', 'P', '', '7B'],
        ],
        col_widths=[16, 12, 30, 24, 22, 14, 25, 14, 8, 10, 10],
        instructions=[
            "Sheet 'Pengguna' mulai baris ke-2 isi data pengguna baru.",
            "",
            "Kolom WAJIB: username, password, nama_lengkap, roles.",
            "Kolom 'roles' diisi salah satu atau gabungan dipisah koma:",
            "  admin, guru, wali_kelas, siswa, tenaga_kependidikan,",
            "  guru_piket, guru_bk, guru_tata_tertib, guru_ekstrakurikuler",
            "",
            "Contoh roles ganda: 'wali_kelas,guru' (tanpa spasi setelah koma).",
            "Kolom 'gender': L atau P (opsional).",
            "Kolom 'kelas_siswa' diisi NAMA kelas (mis. 7A) hanya jika roles berisi 'siswa'.",
            "Kolom 'wali_kelas' diisi NAMA kelas yang dipimpin (mis. 7B) jika roles 'wali_kelas'.",
            "",
            "Sistem akan menolak baris jika username sudah dipakai.",
        ],
    )
    return workbook_to_bytes(wb)


def parse_user_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb['Pengguna'] if 'Pengguna' in wb.sheetnames else wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        rows.append({
            '_row': idx,
            'username': str(row[0]).strip() if row[0] else '',
            'password': str(row[1]).strip() if row[1] else '',
            'full_name': str(row[2]).strip() if row[2] else '',
            'roles': [r.strip() for r in str(row[3]).split(',') if r.strip()] if row[3] else [],
            'nip_nuptk': str(row[4]).strip() if row[4] else None,
            'nisn': str(row[5]).strip() if row[5] else None,
            'email': str(row[6]).strip() if row[6] else None,
            'phone': str(row[7]).strip() if row[7] else None,
            'gender': str(row[8]).strip().upper() if row[8] else None,
            'kelas_siswa': str(row[9]).strip() if row[9] else None,
            'wali_kelas': str(row[10]).strip() if row[10] else None,
        })
    return rows


# ============================================================
# CLASS TEMPLATE
# ============================================================
def class_template() -> bytes:
    wb = _make_workbook_with_data(
        sheet_name="Kelas",
        headers=['nama', 'tingkat', 'paralel', 'wali_kelas_username', 'ruang_kode', 'is_accelerated'],
        examples=[
            ['7A', 7, 'A', 'walas7a', 'R-7A', 'tidak'],
            ['7B', 7, 'B', '', 'R-7B', 'tidak'],
            ['9-AKSEL', 9, 'AKSEL', '', '', 'ya'],
        ],
        col_widths=[12, 10, 10, 22, 12, 14],
        instructions=[
            "Kolom WAJIB: nama, tingkat, paralel.",
            "'tingkat' diisi angka 7, 8, atau 9.",
            "'paralel' = huruf/kode kelas (A, B, AKSEL, dll).",
            "'wali_kelas_username' opsional - username guru wali kelas.",
            "'ruang_kode' opsional - NAMA ruangan default (mis. R-7A).",
            "'is_accelerated' diisi 'ya' atau 'tidak' (default tidak).",
            "",
            "Kelas akan dibuat di Tahun Pelajaran AKTIF saat ini.",
        ],
    )
    return workbook_to_bytes(wb)


def parse_class_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb['Kelas'] if 'Kelas' in wb.sheetnames else wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        try:
            grade = int(row[1]) if row[1] else 7
        except Exception:
            grade = 7
        rows.append({
            '_row': idx,
            'name': str(row[0]).strip() if row[0] else '',
            'grade': grade,
            'parallel': str(row[2]).strip() if row[2] else '',
            'wali_kelas_username': str(row[3]).strip() if row[3] else None,
            'ruang_kode': str(row[4]).strip() if row[4] else None,
            'is_accelerated': str(row[5]).strip().lower() in ('ya', 'yes', 'true', '1') if row[5] else False,
        })
    return rows


# ============================================================
# ROOM TEMPLATE
# ============================================================
def room_template() -> bytes:
    wb = _make_workbook_with_data(
        sheet_name="Ruangan",
        headers=['kode', 'deskripsi', 'lat', 'lon', 'radius_meter', 'gps_aktif', 'qr_mode'],
        examples=[
            ['R-7A', 'Ruang Kelas 7A', -7.9839, 112.6549, 30, 'tidak', 'static'],
            ['R-8B', 'Ruang Kelas 8B', '', '', 20, 'tidak', 'static'],
            ['LAB-IPA', 'Laboratorium IPA', -7.9840, 112.6551, 25, 'ya', 'static'],
        ],
        col_widths=[12, 24, 14, 14, 14, 12, 12],
        instructions=[
            "Kolom WAJIB: kode.",
            "'kode' = identitas ruangan (mis. R-7A, LAB-IPA).",
            "'lat'/'lon' = koordinat GPS (boleh kosong jika tidak pakai GPS).",
            "'radius_meter' = jarak validasi GPS (default 30 m).",
            "'gps_aktif' = 'ya' atau 'tidak' (kalau ya wajib isi lat/lon).",
            "'qr_mode' = 'static' atau 'dynamic'.",
        ],
    )
    return workbook_to_bytes(wb)


def parse_room_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb['Ruangan'] if 'Ruangan' in wb.sheetnames else wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        try:
            lat = float(row[2]) if row[2] not in (None, '') else None
            lon = float(row[3]) if row[3] not in (None, '') else None
            radius = float(row[4]) if row[4] not in (None, '') else 30.0
        except Exception:
            lat, lon, radius = None, None, 30.0
        rows.append({
            '_row': idx,
            'name': str(row[0]).strip() if row[0] else '',
            'description': str(row[1]).strip() if row[1] else None,
            'gps_lat': lat, 'gps_lon': lon,
            'gps_radius_meters': radius,
            'gps_enabled': str(row[5]).strip().lower() in ('ya', 'yes', 'true', '1') if row[5] else False,
            'qr_mode': str(row[6]).strip().lower() if row[6] else 'static',
        })
    return rows


# ============================================================
# SUBJECT TEMPLATE
# ============================================================
def subject_template() -> bytes:
    wb = _make_workbook_with_data(
        sheet_name="MataPelajaran",
        headers=['kode', 'nama'],
        examples=[
            ['MTK', 'Matematika'],
            ['IPA', 'IPA Terpadu'],
            ['BAR', 'Bahasa Arab'],
        ],
        col_widths=[12, 28],
        instructions=[
            "Kolom WAJIB: kode, nama.",
            "'kode' = singkatan unik (mis. MTK, IPA, BIN).",
            "Sistem akan menolak baris jika kode sudah ada di database.",
        ],
    )
    return workbook_to_bytes(wb)


def parse_subject_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb['MataPelajaran'] if 'MataPelajaran' in wb.sheetnames else wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        rows.append({
            '_row': idx,
            'code': str(row[0]).strip().upper() if row[0] else '',
            'name': str(row[1]).strip() if row[1] else '',
        })
    return rows


# ============================================================
# STUDENT TEMPLATE (Siswa)
# ============================================================
def student_template() -> bytes:
    wb = _make_workbook_with_data(
        sheet_name="Siswa",
        headers=['username', 'password', 'nama_lengkap', 'nisn', 'gender', 'kelas',
                 'tempat_lahir', 'tgl_lahir', 'alamat', 'email', 'phone'],
        examples=[
            ['siswa101', 'pwd123', 'Ahmad Faza Pratama', '0098765101', 'L', '7A', 'Malang', '2012-05-10', 'Jl. Cemorokandang 12', '', '081234567000'],
            ['siswa102', 'pwd123', 'Najwa Aulia Rahma', '0098765102', 'P', '7A', 'Malang', '2012-08-22', 'Jl. Sawojajar 5', '', ''],
        ],
        col_widths=[16, 12, 28, 14, 8, 10, 14, 12, 26, 22, 14],
        instructions=[
            "Kolom WAJIB: username, password, nama_lengkap, nisn, kelas.",
            "'kelas' = NAMA kelas seperti terdaftar (mis. 7A, 8B).",
            "'gender': L atau P.",
            "'tgl_lahir' format YYYY-MM-DD (mis. 2012-05-10).",
            "Role siswa akan ditambahkan otomatis.",
        ],
    )
    return workbook_to_bytes(wb)


def parse_student_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb['Siswa'] if 'Siswa' in wb.sheetnames else wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        # Normalize date if datetime
        bd = row[7]
        if hasattr(bd, 'isoformat'):
            try:
                bd = bd.date().isoformat() if hasattr(bd, 'date') else bd.isoformat()
            except Exception:
                bd = str(bd)
        rows.append({
            '_row': idx,
            'username': str(row[0]).strip() if row[0] else '',
            'password': str(row[1]).strip() if row[1] else '',
            'full_name': str(row[2]).strip() if row[2] else '',
            'nisn': str(row[3]).strip() if row[3] else None,
            'gender': str(row[4]).strip().upper() if row[4] else None,
            'kelas': str(row[5]).strip() if row[5] else None,
            'birth_place': str(row[6]).strip() if row[6] else None,
            'birth_date': bd if bd else None,
            'address': str(row[8]).strip() if row[8] else None,
            'email': str(row[9]).strip() if row[9] else None,
            'phone': str(row[10]).strip() if row[10] else None,
        })
    return rows



# ============================================================
# EXPORTERS (snapshot data to Excel)
# ============================================================
def _make_export_workbook(sheet_name: str, headers: List[str],
                          rows: List[List[Any]], col_widths: List[int]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=BRAND_HEX)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thick', color=GOLD_HEX))
    ws.row_dimensions[1].height = 26
    for r in rows:
        ws.append(r)
    for idx, w in enumerate(col_widths):
        letter = ws.cell(row=1, column=idx + 1).column_letter
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = 'A2'
    return workbook_to_bytes(wb)


def export_users_xlsx(users: List[Dict[str, Any]]) -> bytes:
    """Export users (excluding password_hash) to Excel."""
    headers = ['username', 'nama_lengkap', 'nip_nuptk', 'nisn', 'email', 'phone',
               'roles', 'gender', 'kelas_id', 'wali_kelas_id', 'aktif', 'tanggal_dibuat',
               'mutasi', 'tanggal_mutasi']
    rows = []
    for u in users:
        rows.append([
            u.get('username', ''),
            u.get('full_name', ''),
            u.get('nip_nuptk', '') or '',
            u.get('nisn', '') or '',
            u.get('email', '') or '',
            u.get('phone', '') or '',
            ','.join(u.get('roles', [])),
            u.get('gender', '') or '',
            u.get('student_class_id', '') or '',
            u.get('homeroom_class_id', '') or '',
            'Ya' if u.get('is_active', True) else 'Tidak',
            str(u.get('created_at', ''))[:19],
            u.get('mutation_type', '') or '',
            u.get('mutation_date', '') or '',
        ])
    widths = [16, 28, 16, 14, 26, 16, 22, 8, 18, 18, 8, 20, 10, 14]
    return _make_export_workbook('Users', headers, rows, widths)


def export_students_xlsx(students: List[Dict[str, Any]]) -> bytes:
    """Export students with class name resolved."""
    headers = ['username', 'NISN', 'nama_lengkap', 'kelas', 'gender',
               'tempat_lahir', 'tanggal_lahir', 'alamat', 'email', 'phone',
               'aktif', 'mutasi', 'tanggal_mutasi']
    rows = []
    for s in students:
        rows.append([
            s.get('username', ''),
            s.get('nisn', '') or '',
            s.get('full_name', ''),
            s.get('class_name', '') or '',
            s.get('gender', '') or '',
            s.get('birth_place', '') or '',
            s.get('birth_date', '') or '',
            s.get('address', '') or '',
            s.get('email', '') or '',
            s.get('phone', '') or '',
            'Ya' if s.get('is_active', True) else 'Tidak',
            s.get('mutation_type', '') or '',
            s.get('mutation_date', '') or '',
        ])
    widths = [16, 14, 28, 10, 8, 20, 14, 30, 26, 16, 8, 10, 14]
    return _make_export_workbook('Siswa', headers, rows, widths)


def export_schedules_xlsx(schedules: List[Dict[str, Any]]) -> bytes:
    """Export schedules with resolved names."""
    headers = ['hari', 'jam_mulai', 'jam_selesai', 'kelas', 'mapel_kode', 'mapel',
               'guru', 'ruang', 'semester', 'status']
    rows = []
    for s in schedules:
        rows.append([
            s.get('day', ''),
            s.get('start_time', ''),
            s.get('end_time', ''),
            s.get('class_name', '') or '',
            s.get('subject_code', '') or '',
            s.get('subject_name', '') or '',
            s.get('teacher_name', '') or '',
            s.get('room_name', '') or '',
            s.get('semester', '') or '',
            s.get('status', 'submitted'),
        ])
    widths = [10, 10, 10, 10, 10, 22, 22, 12, 10, 12]
    return _make_export_workbook('Jadwal', headers, rows, widths)


def export_grades_xlsx(grades: List[Dict[str, Any]]) -> bytes:
    """Export grade entries."""
    headers = ['NISN', 'nama_siswa', 'kelas', 'mapel_kode', 'mapel',
               'semester', 'nilai_pengetahuan', 'nilai_keterampilan', 'nilai_akhir',
               'predikat', 'deskripsi', 'tanggal_input']
    rows = []
    for g in grades:
        rows.append([
            g.get('student_nisn', '') or '',
            g.get('student_name', '') or '',
            g.get('class_name', '') or '',
            g.get('subject_code', '') or '',
            g.get('subject_name', '') or '',
            g.get('semester', '') or '',
            g.get('nilai_pengetahuan', '') if g.get('nilai_pengetahuan') is not None else '',
            g.get('nilai_keterampilan', '') if g.get('nilai_keterampilan') is not None else '',
            g.get('nilai_akhir', '') if g.get('nilai_akhir') is not None else '',
            g.get('predicate', '') or '',
            g.get('description', '') or '',
            str(g.get('submitted_at', ''))[:19],
        ])
    widths = [14, 28, 10, 12, 22, 10, 14, 14, 12, 10, 30, 20]
    return _make_export_workbook('Nilai', headers, rows, widths)


# ============================================================
# TEMPLATE UPDATE NISM & NOMOR PESERTA UJIAN
# ============================================================
def nism_update_template(students: List[Dict[str, Any]]) -> bytes:
    """
    Generate template Excel untuk update NISM dan Nomor Peserta Ujian Madrasah.
    Template berisi data siswa existing yang bisa diupdate.
    """
    headers = ['id', 'nisn', 'nama_lengkap', 'kelas', 'nism', 'nomor_peserta_ujian']
    examples = []

    # Populate dengan data siswa yang ada
    for s in students:
        examples.append([
            s.get('id', ''),
            s.get('nisn', ''),
            s.get('full_name', ''),
            s.get('class_name', ''),
            s.get('nism', ''),
            s.get('nomor_peserta_ujian', ''),
        ])

    col_widths = [36, 14, 30, 10, 16, 20]
    instructions = [
        "Sheet 'Update NISM & No. Ujian' berisi data siswa yang sudah ada.",
        "",
        "Cara Penggunaan:",
        "1. Download template ini untuk mendapatkan data siswa terkini",
        "2. Isi kolom 'nism' dan 'nomor_peserta_ujian' sesuai kebutuhan",
        "3. JANGAN UBAH kolom 'id', 'nisn', 'nama_lengkap', dan 'kelas' (untuk referensi saja)",
        "4. Upload kembali file ini untuk update data",
        "",
        "Kolom yang bisa diupdate:",
        "- nism: NIS Madrasah (NISM) - diperlukan untuk buku induk",
        "- nomor_peserta_ujian: Nomor Peserta Ujian Madrasah (untuk siswa kelas 9)",
        "",
        "Kolom yang TIDAK boleh diubah (hanya untuk referensi):",
        "- id: ID sistem (wajib ada, jangan diubah)",
        "- nisn: Nomor Induk Siswa Nasional (referensi)",
        "- nama_lengkap: Nama siswa (referensi)",
        "- kelas: Kelas siswa (referensi)",
        "",
        "Tips:",
        "- Kosongkan kolom jika tidak ingin mengubah nilai",
        "- Sistem hanya akan update baris yang memiliki nilai baru di kolom nism atau nomor_peserta_ujian",
    ]

    wb = _make_workbook_with_data(
        sheet_name="Update NISM & No. Ujian",
        headers=headers,
        examples=examples,
        col_widths=col_widths,
        instructions=instructions
    )

    return workbook_to_bytes(wb)


def parse_nism_update_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parse Excel file untuk update NISM dan Nomor Peserta Ujian.
    Returns list of dicts dengan keys: id, nism, nomor_peserta_ujian, _row
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []

    # Headers expected: id, nisn, nama_lengkap, kelas, nism, nomor_peserta_ujian
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):  # Skip empty rows
            continue

        # row[0] = id, row[4] = nism, row[5] = nomor_peserta_ujian
        student_id = str(row[0]).strip() if row[0] else ''
        nism = str(row[4]).strip() if row[4] and str(row[4]).strip() else None
        nomor_ujian = str(row[5]).strip() if row[5] and str(row[5]).strip() else None

        # Skip jika tidak ada ID (mandatory)
        if not student_id:
            continue

        # Skip jika tidak ada perubahan (nism dan nomor_ujian kosong)
        if not nism and not nomor_ujian:
            continue

        rows.append({
            'id': student_id,
            'nism': nism,
            'nomor_peserta_ujian': nomor_ujian,
            '_row': idx
        })

    return rows
